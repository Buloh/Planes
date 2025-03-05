# ----- Importy a globÃ¡lnÃ­ konstanty
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog
import logging
import socket
import sqlite3
from openpyxl import load_workbook
import json
import calendar
from datetime import datetime, date, timedelta
import base64
import os
import time
import getpass

# ----- Verze programu
__version__="1.2.0"

# -----1.0.0 ZÃ¡kladnÃ­ funkÄnost programu: NÃ¡hled, Editace, Role, PÅ™Ã­stupy, Logy, ZÃ¡kladnÃ­ nastavenÃ­, VytvoÅ™enÃ­ databÃ¡ze, UloÅ¾enÃ­ zÃ¡kladnÃ­ho nastavenÃ­, Import dat z Ecxelu
# ----- 1.1.0 DoplnÄ›nÃ­ o funkci: ZamknutÃ­ PlÃ¡nu a OdeknutÃ­ plÃ¡nu s ukÃ¡dÃ¡nÃ­m do GlobÃ¡lnÃ­ho nastavenÃ­ Pro Admina a Superadmina
# ------ 1.2.0 PÅ™idÃ¡na fuknce zavÅ™enÃ­ okna po odhlÃ¡Å¡enÃ­ uÅ¾ivatele

# PÅ™idejte funkci show_verze() na globÃ¡lnÃ­ ÃºroveÅˆ
def show_verze():
    info_text = "SprÃ¡va PlÃ¡nu SluÅ¾eb a DovolenÃ½ch \nVerze: " + __version__ + "\n(C) 2025 Holub Stanislav"
    messagebox.showinfo("O aplikaci", info_text)

# ----- Mapa barev: pÅ™evod ÄeskÃ½ch nÃ¡zvÅ¯ na anglickÃ©
COLOR_MAP = {
    "Å¾lutÃ©": "yellow",
    "zelenÃ©": "green",
    "oranÅ¾ovÃ©": "orange",
    "rÅ¯Å¾ovÃ©": "pink",
    "modrÃ©": "blue",
    "hnÄ›dÃ©": "brown",
    "neutrÃ¡lnÃ­": "SystemButtonFace",
    "fialovÃ©": "purple",
    "tyrkysovÃ©": "turquoise",
    "Å¡edÃ©": "gray"
}

# PÅ™idejte funkci show_help() na globÃ¡lnÃ­ ÃºroveÅˆ
def show_help():
    """OtevÅ™e okno s rozÅ¡Ã­Å™enou nÃ¡povÄ›dou aplikace podle rolÃ­."""
    help_win = tk.Toplevel(root)
    help_win.title("NÃ¡povÄ›da")
    help_win.geometry("700x550")
    
    help_text = (
        "SprÃ¡va PlÃ¡nu SluÅ¾eb a DovolenÃ½ch verze: "+__version__+"\n\n"
        "Tento program slouÅ¾Ã­ k nÃ¡vrhu plÃ¡nu sluÅ¾eb, dovolenÃ½ch a vypuÅ¡tÄ›nÃ½ch smÄ›na. Aplikace nabÃ­zÃ­ rÅ¯znÃ© funkcen\n"
        "pro zadÃ¡vÃ¡nÃ­, editaci a sprÃ¡vu plÃ¡nÅ¯, a to prostÅ™ednictvÃ­m pÅ™ehlednÃ©ho grafickÃ©ho rozhranÃ­.\n\n"
        
        "OMEZENÃ ODPOVÄšDNOSTI A UÅ½ÃVÃNÃ NA VLASTNÃ NEBEZPEÄŒÃ:\n\n"
        "PouÅ¾Ã­vÃ¡nÃ­ tÃ©to aplikace probÃ­hÃ¡ na vlastnÃ­ odpovÄ›dnost uÅ¾ivatele. Autor, vÃ½vojÃ¡Å™i a distributoÅ™i tohoto\n" 
        "softwaru neposkytujÃ­ Å¾Ã¡dnÃ© zÃ¡ruky, a to ani v rozsahu vÃ½slovnÃ©m, ani pÅ™edpoklÃ¡danÃ©m, vÄetnÄ› zÃ¡ruk obchodovatelnosti\n"
        "Äi vhodnosti pro konkrÃ©tnÃ­ ÃºÄely. V Å¾Ã¡dnÃ©m pÅ™Ã­padÄ› nebudou odpovÄ›dnÃ­ za jakÃ©koliv pÅ™Ã­mÃ©, nepÅ™Ã­mÃ©, nÃ¡hodnÃ© Äi\n"
        "nÃ¡slednÃ© Å¡kody, ztrÃ¡tu dat nebo jinÃ© nepÅ™Ã­znivÃ© dÅ¯sledky vyplÃ½vajÃ­cÃ­ z pouÅ¾Ã­vÃ¡nÃ­ aplikace. UÅ¾ivatelÃ© jsou\n"
        "povinni ovÄ›Å™it sprÃ¡vnost a Ãºplnost vÅ¡ech dat pÅ™ed jejich pouÅ¾itÃ­m.\n\n"

        "RYCHLÃ NÃPOVÄšDA\n\n"
        "1. VÃ½bÄ›r zÃ¡loÅ¾ky:\n"
        "- pÅ™ejdi na zÃ¡loÅ¾ku ZAMÄšSTANEC pro plÃ¡n na celÃ½ rok, nebo na zÃ¡loÅ¾ku SmÄ›ny pro plÃ¡n na zvolenÃ½ mÄ›sÃ­c.\n\n"

        "2. ZÃ¡loÅ¾ka ZamÄ›stnanec:\n"
        "VÃ½bÄ›r zamÄ›stnance a roku: V zÃ¡loÅ¾ce ZamÄ›stnanec vyberte svÃ© jmÃ©no ze seznamu a zvolte poÅ¾adovanÃ½ rok.\n"
        "ZobrazenÃ­ plÃ¡nu: KliknÄ›te na tlaÄÃ­tko Zobrazit plÃ¡n, ÄÃ­mÅ¾ se naÄte vÃ¡Å¡ detailnÃ­ plÃ¡n smÄ›n.\n\n"

        "3. ZÃ¡loÅ¾ka SmÄ›ny:\n"
        "VÃ½bÄ›r filtrÅ¯: PÅ™ejdÄ›te do zÃ¡loÅ¾ky SmÄ›ny.\n"
        "NastavenÃ­ filtrÅ¯: Vyberte z rozevÃ­racÃ­ch seznamÅ¯ rok, mÄ›sÃ­c a konkrÃ©tnÃ­ smÄ›nu, pro kterou chcete plÃ¡n zobrazit.\n"
        "NaÄtenÃ­ plÃ¡nu: KliknÄ›te na tlaÄÃ­tko Zobraz plÃ¡n SmÄ›ny, kterÃ© naÄte plÃ¡ny pro zadanÃ¡ kritÃ©ria.\n\n"

        "Tento postup vÃ¡m umoÅ¾nÃ­ rychle najÃ­t a zobrazit svÅ¯j plÃ¡n smÄ›n, aÅ¥ uÅ¾ hledÃ¡te pod svou osobou nebo chcete prohlÃ­Å¾et plÃ¡ny celÃ© smÄ›ny.\n\n\n"

        "PODROBNÄšJÅ Ã NÃVOD\n\n"

        "Role v aplikaci:\n"
        "   - UÅ¾ivatel: MÃ¡ omezenÃ½ pÅ™Ã­stup k prohlÃ­Å¾enÃ­ a zÃ¡kladnÃ­mu filtrovÃ¡nÃ­ plÃ¡nÅ¯.\n"
        "   - Velitel: MÅ¯Å¾e upravovat plÃ¡ny svÃ© smÄ›ny a mÃ¡ rozÅ¡Ã­Å™enÃ¡ prÃ¡va pro zadÃ¡vÃ¡nÃ­ a kontrolu smÄ›n a dovolenÃ½ch.\n"
        "   - Admin: MÃ¡ plnÃ¡ oprÃ¡vnÄ›nÃ­ k ÃºpravÃ¡m, sprÃ¡vÄ› databÃ¡ze a konfiguraci aplikace.\n"
        "   - Superadmin: MÃ¡ nejvyÅ¡Å¡Ã­ oprÃ¡vnÄ›nÃ­, vÄetnÄ› zÃ¡sahu do globÃ¡lnÃ­ch nastavenÃ­ a ÃºdrÅ¾by systÃ©mu.\n\n"
        
        "NÃ¡povÄ›da aplikace:\n\n"
        "1. ZadÃ¡vÃ¡nÃ­ a editace plÃ¡nÅ¯:\n"
        "   - Aplikace umoÅ¾Åˆuje zadÃ¡vat a editovat plÃ¡ny sluÅ¾eb pomocÃ­ pÅ™ehlednÃ©ho grafickÃ©ho rozhranÃ­.\n"
        "   - PlÃ¡ny lze editovat pÅ™Ã­mo v tabulkÃ¡ch, kde jsou zobrazeny podrobnosti o jednotlivÃ½ch dnech.\n\n"
        
        "2. LogovÃ¡nÃ­ a bezpeÄnost:\n"
        "   - KaÅ¾dÃ¡ akce (pÅ™ihlÃ¡Å¡enÃ­, Ãºpravy, mazÃ¡nÃ­) je automaticky zaznamenÃ¡na do log souboru.\n"
        "   - PÅ™Ã­stupovÃ¡ prÃ¡va jsou nastavena tak, Å¾e kaÅ¾dÃ½ uÅ¾ivatel mÃ¡ pÅ™Ã­stup pouze k funkcÃ­m odpovÃ­dajÃ­cÃ­m jeho roli.\n\n"
        
        "3. NastavenÃ­ a konfigurace:\n"
        "   - V sekci 'NastavenÃ­' jen superadmin mÅ¯Å¾ete mÄ›nit globÃ¡lnÃ­ parametry, jako jsou Ãºvazky, smÄ›ny, barvy a poÄet hodin pÅ™iÅ™azenÃ½ch jednotlivÃ½m smÄ›nÃ¡m.\n"
        "   - Tyto hodnoty ovlivÅˆujÃ­ vÃ½poÄty v aplikaci, napÅ™Ã­klad celkovÃ½ souÄet hodin a poÄet smÄ›n v plÃ¡novanÃ½ch sluÅ¾bÃ¡ch.\n\n"
        
        "4. DalÅ¡Ã­ funkce:\n"
        "   - Aplikace umoÅ¾Åˆuje naÄÃ­tÃ¡nÃ­ dat z Excelu, automatickÃ© zÃ¡lohovÃ¡nÃ­ databÃ¡ze a filtrovÃ¡nÃ­ zÃ¡znamÅ¯ podle rÅ¯znÃ½ch kritÃ©riÃ­.\n\n"
        
        "SpecifickÃ© informace podle rolÃ­.\n\n"
        
        "   UÅ¾ivatel:\n"
        "      - MÅ¯Å¾e pouze prohlÃ­Å¾et plÃ¡ny a vyuÅ¾Ã­vat zÃ¡kladnÃ­ filtrovÃ¡nÃ­.\n"
        "      - NemÃ¡ oprÃ¡vnÄ›nÃ­ k ÃºpravÃ¡m nebo mazÃ¡nÃ­ zÃ¡znamÅ¯.\n\n"
        
        "   Velitel:\n"
        "      - MÃ¡ rozÅ¡Ã­Å™enÃ¡ oprÃ¡vnÄ›nÃ­ k zadÃ¡vÃ¡nÃ­ a ÃºpravÄ› VlastnÃ­ch plÃ¡nÅ¯ smÄ›ny.\n"
        "      - MÅ¯Å¾e kontrolovat a upravovat svÃ© smÄ›ny, ale nemÅ¯Å¾e zasahovat do plÃ¡nÅ¯ ostatnÃ­ch.\n\n"
        
        "   Admin:\n"
        "      - MÃ¡ plnÃ¡ prÃ¡va k editaci, SprÃ¡vÄ› databÃ¡ze a Konfiguraci aplikace.\n"
        "      - MÅ¯Å¾e mÄ›nit GlobÃ¡lnÃ­ nastavenÃ­ a zadÃ¡vat novÃ© Ãºdaje, kterÃ© ovlivÅˆujÃ­ vÅ¡echny plÃ¡ny.\n"
        "      - MÅ¯Å¾e uzamknout plÃ¡n sluÅ¾eb, po uazmÄenÃ­ nejde plÃ¡n editovat.\n\n"
        
        "   Superadmin:\n"
        "      - MÃ¡ nejvyÅ¡Å¡Ã­ oprÃ¡vnÄ›nÃ­ v aplikaci.\n"
        "      - MÅ¯Å¾e provÃ¡dÄ›t zÃ¡sahy do celÃ©ho systÃ©mu, vÄetnÄ› Ãºprav GlobÃ¡lnÃ­ch nastavenÃ­, SprÃ¡vy logÅ¯ a DatabÃ¡ze.\n"
        "      - MÅ¯Å¾e uzamknout plÃ¡n sluÅ¾eb, po uazmÄenÃ­ nejde plÃ¡n editovat.\n\n"
        
        "5. UÅ¾ivatelskÃ¡ podpora:\n"
        "   - Pro dalÅ¡Ã­ informace nebo Å™eÅ¡enÃ­ problÃ©mÅ¯ kontaktujte sprÃ¡vce programu. Pokud zjistÃ­ chybu nahlaÅ¡te ji sprÃ¡vci.\n\n"
        
        "Tento nÃ¡vod shrnuje hlavnÃ­ funkce a principy aplikace. Pro detailnÄ›jÅ¡Ã­ informace kontaktujte sprÃ¡vce programu.\n\n"
        
        "VysvÄ›tlivky k informaÄnÃ­ tabulce u zamÄ›stnacÅ¯:\n"
        "   - NovÃ¡ dovolenÃ¡: PoÄet hodin aktuÃ¡lnÄ› pÅ™iÄtenÃ½ch jako dovolenÃ¡ v danÃ©m roce.\n"
        "   - StarÃ¡ dovolenÃ¡: PoÄet hodin dovolenÃ© pÅ™enesenÃ½ch z pÅ™edchozÃ­ho obdobÃ­.\n"
        "   - Celkem dovolenÃ¡: SouÄet novÃ© a starÃ© dovolenÃ©.\n"
        "   - NaplÃ¡novat Dov: PoÄet hodin dovolenÃ©, kterÃ© jsou jiÅ¾ naplÃ¡novÃ¡ny v rozvrhu.\n"
        "   - RozdÃ­l PlÃ¡n a NÃ¡rok: RozdÃ­l mezi naplÃ¡novanou dovolenou v rozvrhu a skuteÄnÃ½m nÃ¡rokem na dovolenou.\n"
        "     ( nÄ›mÄ›la by se zde ukÃ¡zat zÃ¡pornÃ¡ hodnota !!!)\n"
        "   - Celkem smÄ›n: CelkovÃ½ poÄet smÄ›n zaznamenanÃ½ch v plÃ¡nu.\n\n"
        "     K VYROVNÃNÃ HODIN V KALENDÃÅ˜NÃM ROCE:\n"
        "   - Klouz I.: PoÄet vypuÅ¡tÄ›nÃ½ch smÄ›na za  I. pololetÃ­. (X/X)\n"
        "     (X naplÃ¡novat v I. pololetÃ­ / X naplÃ¡novÃ¡no I. pololetÃ­ )\n\n"
        "   - Klouz II.: PoÄet vypuÅ¡tÄ›nÃ½ch smÄ›na za II. pololetÃ­.\n"
        "     (X naplÃ¡novat v II. pololetÃ­ / X naplÃ¡novÃ¡no II. pololetÃ­)\n\n"     
        "   - r I.: PoÄet rannÃ­ch smÄ›n v I. pololetÃ­.\n"
        "     (r naplÃ¡novat v I. pololetÃ­ / r naplÃ¡novÃ¡no I. pololetÃ­ )\n\n"
        "   - r II.: PoÄet rannÃ­ch smÄ›n ve I. pololetÃ­.\n"
        "     (r naplÃ¡novat v II. pololetÃ­ / r naplÃ¡novÃ¡no II. pololetÃ­ )\n\n"
        "   - zkratky v programu se shodujÃ­ s tiÅ¡tÄ›nou verzÃ­ PlÃ¡nu sluÅ¾eb\n\n"

        "Licence pro otevÅ™enÃ½ software\n\n"

        "Copyright (c) [2025] [Holub Stanislav]\n"
        "Tento software je poskytovÃ¡n pod licencÃ­ pro otevÅ™enÃ½ software, dÃ¡le jen â€Licenceâ€œ. UÅ¾ivatelÃ© tÃ©to licence majÃ­\n" 
        "prÃ¡vo pouÅ¾Ã­vat, kopÃ­rovat, upravovat, spojovat, publikovat, distribuovat, sublicencovat a/nebo prodÃ¡vat kopie softwaru,\n"
        "za nÃ¡sledujÃ­cÃ­ch podmÃ­nek:\n\n"
        "1. Tento software je poskytovÃ¡n â€jak jeâ€œ, bez jakÃ½chkoli zÃ¡ruk, vyjÃ¡dÅ™enÃ½ch nebo implicitnÃ­ch, vÄetnÄ›, ale neomezujÃ­cÃ­ se\n"
        "na implicitnÃ­ zÃ¡ruky obchodovatelnosti, vhodnosti pro urÄitÃ½ ÃºÄel a nezÃ¡nÄ›tÃ½m prÃ¡vÅ¯m. AutoÅ™i nebo vlastnÃ­ci autorskÃ½ch prÃ¡v\n" 
        "nejsou zodpovÄ›dnÃ­ za jakÃ©koli nÃ¡roky, Å¡kody nebo jinou odpovÄ›dnost, aÅ¥ uÅ¾ ve smlouvÄ›, nebo z jinÃ©ho dÅ¯vodu, plynoucÃ­ z nebo\n" 
        "v souvislosti se softwarem nebo jeho pouÅ¾Ã­vÃ¡nÃ­m.\n\n"
        "2. JmÃ©na â€[vlastnÃ­k autorskÃ½ch prÃ¡v]â€œ nemohou bÃ½t pouÅ¾ity k podpoÅ™e nebo propagaci vÃ½robkÅ¯ odvozenÃ½ch z tohoto softwaru\n" 
        "bez pÅ™edchozÃ­ho pÃ­semnÃ©ho povolenÃ­.\n\n"
    )
    
    text_widget = tk.Text(help_win, wrap="word", font=("TkDefaultFont", 10))
    text_widget.insert("1.0", help_text)
    text_widget.config(state="disabled")
    text_widget.pack(expand=True, fill="both", padx=10, pady=10)
    
    tk.Button(help_win, text="ZavÅ™Ã­t", command=help_win.destroy).pack(pady=5)

# NÃ¡zev konfiguraÄnÃ­ho souboru
CONFIG_FILE = "global_settings.json"

# ----- Funkce pro zakÃ³dovÃ¡nÃ­ a dekÃ³dovÃ¡nÃ­ hesla
def encode_password(pwd):
    """ZakÃ³duje heslo pomocÃ­ base64."""
    return base64.b64encode(pwd.encode("utf-8")).decode("utf-8")

def decode_password(encoded_pwd):
    """DekÃ³duje heslo zakÃ³dovanÃ© pomocÃ­ base64."""
    return base64.b64decode(encoded_pwd.encode("utf-8")).decode("utf-8")

# ----- UklÃ¡dÃ¡nÃ­ a naÄÃ­tÃ¡nÃ­ konfigurace
def save_config(config):
    """UloÅ¾Ã­ nastavenÃ­ do konfiguraÄnÃ­ho souboru."""
    try:
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump(config, f, ensure_ascii=False, indent=4)
    except Exception as e:
        logging.error(f"Chyba pÅ™i uklÃ¡dÃ¡nÃ­ konfiguraÄnÃ­ho souboru: {e}")
        messagebox.showerror("Chyba", f"Chyba pÅ™i uklÃ¡dÃ¡nÃ­ nastavenÃ­: {e}")

def load_config():
    """
    NaÄte konfiguraÄnÃ­ soubor.
    Pokud soubor neexistuje, vytvoÅ™Ã­ vÃ½chozÃ­ konfiguraci s pÅ™Ã­stupovÃ½mi zÃ¡znamy.
    Pokud soubor existuje, ale chybÃ­ v nÄ›m zÃ¡znam pro superadmin, doplnÃ­ jej.
    """
    try:
        if os.path.exists(CONFIG_FILE):
            with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                config = json.load(f)
            # ----- Kontrola existence zÃ¡znamu pro superadmin
            access_list = config.get("access", [])
            if not any(entry[2] == "superadmin" for entry in access_list):
                access_list.insert(0, ("superadmin", encode_password("12345"), "superadmin"))
                config["access"] = access_list
                save_config(config)
        else:
            # VÃ½chozÃ­ konfigurace, pokud soubor neexistuje
            config = {
                "40": [],
                "37.5": [],
                "37.75": [],
                "access": [
                    ("superadmin", encode_password("12345"), "superadmin"),
                    ("Velitel smÄ›ny 1", encode_password("heslo1"), "velitel"),
                    ("Velitel smÄ›ny 2", encode_password("heslo2"), "velitel"),
                    ("Velitel smÄ›ny 3", encode_password("heslo3"), "velitel"),
                    ("Velitel smÄ›ny 4", encode_password("heslo4"), "velitel"),
                    ("Velitel smÄ›ny 5", encode_password("heslo5"), "velitel"),
                    ("Velitel smÄ›ny 6", encode_password("heslo6"), "velitel"),
                ],
                "locked_plans": {}  # MOD: VÃ½chozÃ­ prÃ¡zdnÃ½ slovnÃ­k pro zÃ¡mky plÃ¡nÅ¯
            }
            save_config(config)
        # ZajiÅ¡tÄ›nÃ­, Å¾e vÅ¡echny klÃ­Äe existujÃ­
        for key in ["40", "37.5", "37.75", "access"]:
            if key not in config:
                config[key] = [] if key != "access" else []
        if "locked_plans" not in config:
            config["locked_plans"] = {}  # MOD: UjistÃ­me se, Å¾e existuje i klÃ­Ä pro zÃ¡mky
        return config
    except Exception as e:
        logging.error(f"Chyba pÅ™i naÄÃ­tÃ¡nÃ­ konfiguraÄnÃ­ho souboru: {e}")
        messagebox.showerror("Chyba", f"Chyba pÅ™i naÄÃ­tÃ¡nÃ­ nastavenÃ­: {e}")
        return {}

global_settings = load_config()

# ----- GlobÃ¡lnÃ­ promÄ›nnÃ© pro sprÃ¡vu uÅ¾ivatele a plÃ¡nÅ¯
current_user_role = "uÅ¾ivatel"      # Role aktuÃ¡lnÄ› pÅ™ihlÃ¡Å¡enÃ©ho uÅ¾ivatele
current_user_name = "uÅ¾ivatel"      # JmÃ©no aktuÃ¡lnÄ› pÅ™ihlÃ¡Å¡enÃ©ho uÅ¾ivatele
current_user_shift = None           # Pouze pro filtrovÃ¡nÃ­, nikoli pro oprÃ¡vnÄ›nÃ­
current_record_id = None            # UchovÃ¡vÃ¡ id aktuÃ¡lnÄ› zobrazenÃ©ho plÃ¡nu
month_frames = {}                   # SlovnÃ­k pro uloÅ¾enÃ­ odkazÅ¯ na jednotlivÃ© mÄ›sÃ­ÄnÃ­ rÃ¡mce

# Deklarace globÃ¡lnÃ­ch promÄ›nnÃ½ch pro tlaÄÃ­tka zÃ¡mku plÃ¡nu
btn_lock_plan = None
btn_unlock_plan = None

# ----- SÃ­Å¥ovÃ© a logovacÃ­ funkce
def get_ip_address():
    """
    VracÃ­ aktuÃ¡lnÃ­ IP adresu.
    PouÅ¾Ã­vÃ¡ se pÅ™i logovÃ¡nÃ­ akcÃ­.
    """
    try:
        with socket.socket(socket.AF_INET, socket.SOCK_DGRAM) as s:
            s.connect(("8.8.8.8", 80))
            ip = s.getsockname()[0]
    except Exception:
        try:
            ip = socket.gethostbyname(socket.gethostname())
            if ip.startswith("127."):
                addresses = [addr[4][0] for addr in socket.getaddrinfo(socket.gethostname(), None)
                             if not addr[4][0].startswith("127.")]
                ip = addresses[0] if addresses else "N/A"
        except Exception:
            ip = "N/A"
    return ip

def log_action(action):
    """
    ZapÃ­Å¡e akci do logu, vÄetnÄ› informace o IP adrese a roli uÅ¾ivatele.
    """
    ip = get_ip_address()
    logged_in_user = getpass.getuser()  # nebo mÅ¯Å¾ete pouÅ¾Ã­t os.getlogin(
    role = current_user_role if current_user_role else "neznÃ¡mÃ¡"
    message = f"{action} - Role: {role} - IP: {ip} - PC uÅ¾ivatel: {logged_in_user}"
    logging.info(message)

logging.basicConfig(
    filename="service_plan_log.txt",
    level=logging.INFO,
    format='%(asctime)s - %(message)s',
    datefmt='%d.%m.%Y %H:%M:%S'
)

# ----- Funkce pro zÃ¡lohovÃ¡nÃ­ databÃ¡ze
def backup_database():
    """
    VytvoÅ™Ã­ zÃ¡lohu databÃ¡ze do adresÃ¡Å™e 'db_backups'.
    NÃ¡zev zÃ¡lohy obsahuje datum a Äas.
    """
    try:
        backup_dir = "db_backups"
        if not os.path.exists(backup_dir):
            os.makedirs(backup_dir)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_filename = f"service_plans_backup_{timestamp}.db"
        backup_path = os.path.join(backup_dir, backup_filename)
        with sqlite3.connect("service_plans.db") as src:
            with sqlite3.connect(backup_path) as dest:
                src.backup(dest)
        log_action("ZÃ¡loha databÃ¡ze byla ÃºspÄ›Å¡nÄ› vytvoÅ™ena")
    except Exception as e:
        logging.error(f"Chyba pÅ™i zÃ¡loze databÃ¡ze: {e}")
        messagebox.showerror("Chyba", f"Chyba pÅ™i zÃ¡loze databÃ¡ze: {e}")

# ----- Inicializace databÃ¡ze a vytvoÅ™enÃ­ tabulky
def init_db():
    """
    Inicializuje databÃ¡zi a vytvoÅ™Ã­ tabulku 'plans', pokud jeÅ¡tÄ› neexistuje.
    """
    try:
        with sqlite3.connect("service_plans.db") as conn:
            conn.execute("PRAGMA foreign_keys = ON")
            conn.execute("""
                CREATE TABLE IF NOT EXISTS plans (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    osobni_cislo TEXT,
                    jmeno_prijmeni TEXT,
                    smena TEXT,
                    uvazek TEXT,
                    roky TEXT,
                    poradi TEXT,
                    stara_dovolena TEXT DEFAULT "0",
                    dovolena TEXT DEFAULT "0",
                    leden TEXT,
                    unor TEXT,
                    brezen TEXT,
                    duben TEXT,
                    kveten TEXT,
                    cerven TEXT,
                    cervenec TEXT,
                    srpen TEXT,
                    zari TEXT,
                    rijen TEXT,
                    listopad TEXT,
                    prosinec TEXT
                )
            """)
            conn.commit()
    except Exception as e:
        logging.error(f"Chyba pÅ™i inicializaci databÃ¡ze: {e}")
        messagebox.showerror("Chyba", f"Chyba pÅ™i inicializaci databÃ¡ze: {e}")

# ----- Funkce pro aktualizaci Treeview (filtrace a zobrazenÃ­ dat)
def refresh_treeview_filtered(jmeno_filter="", rok_filter="", smena_filter=""):
    """
    Aktualizuje zobrazenÃ­ seznamu zÃ¡znamÅ¯ v Treeview podle zadanÃ½ch filtrÅ¯.
    """
    try:
        for item in tree.get_children():
            tree.delete(item)
        query = "SELECT id, jmeno_prijmeni, smena, uvazek, roky FROM plans WHERE 1=1"
        params = []
        if jmeno_filter:
            query += " AND jmeno_prijmeni = ?"
            params.append(jmeno_filter)
        if rok_filter:
            query += " AND roky = ?"
            params.append(rok_filter)
        if smena_filter != "":
            query += " AND smena = ?"
            params.append(smena_filter)
        with sqlite3.connect("service_plans.db") as conn:
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            cursor.execute(query, params)
            for row in cursor.fetchall():
                tree.insert("", tk.END, iid=row["id"],
                            values=(row["jmeno_prijmeni"], row["smena"], row["uvazek"], row["roky"]))
    except Exception as e:
        logging.error(f"Chyba pÅ™i naÄÃ­tÃ¡nÃ­ dat do Treeview: {e}")
        messagebox.showerror("Chyba", f"Chyba pÅ™i naÄÃ­tÃ¡nÃ­ dat: {e}")

def refresh_treeview():
    """ObnovÃ­ Treeview a resetuje filtry."""
    filter_jmeno.set("")
    filter_rok.set(str(datetime.now().year))
    filter_smena.set("")
    filter_smena.config(state="readonly")
    refresh_treeview_filtered()

# ----- PomocnÃ© funkce a vÃ½poÄty
def compute_easter(year):
    """
    VypoÄÃ­tÃ¡ datum Velikonoc pro danÃ½ rok.
    """
    a = year % 19
    b = year // 100
    c = year % 100
    d = b // 4
    e = b % 4
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19 * a + b - d - g + 15) % 30
    i = c // 4
    k = c % 4
    l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451
    month = (h + l - 7 * m + 114) // 31
    day = ((h + l - 7 * m + 114) % 31) + 1
    return date(year, month, day)

def load_month_data(ws, month_rows):
    """
    NaÄte data z ExcelovÃ©ho listu pro danÃ© Å™Ã¡dky odpovÃ­dajÃ­cÃ­ jednotlivÃ½m mÄ›sÃ­cÅ¯m.
    """
    try:
        data = {}
        for month, row in month_rows.items():
            row_values = [cell for cell in next(ws.iter_rows(min_row=row, max_row=row,
                                                               min_col=1, max_col=32, values_only=True))]
            data[month] = json.dumps(row_values, ensure_ascii=False)
        return data
    except Exception as e:
        logging.error(f"Chyba pÅ™i naÄÃ­tÃ¡nÃ­ dat z Excelu: {e}")
        messagebox.showerror("Chyba", f"Chyba pÅ™i naÄÃ­tÃ¡nÃ­ dat z Excelu: {e}")
        return {}

def treeview_sort_column(tv, col, reverse):
    """
    SeÅ™adÃ­ sloupec v Treeview podle hodnot.
    """
    try:
        l = [(tv.set(k, col), k) for k in tv.get_children('')]
        l.sort(key=lambda t: t[0], reverse=reverse)
        for index, (val, k) in enumerate(l):
            tv.move(k, '', index)
        tv.heading(col, command=lambda: treeview_sort_column(tv, col, not reverse))
    except Exception as e:
        logging.error(f"Chyba pÅ™i Å™azenÃ­ sloupce {col}: {e}")

def calculate_month_summary(day_plan_list, uvazek):
    """
    VypoÄÃ­tÃ¡ souÄet hodin a Äetnost jednotlivÃ½ch smÄ›n pro danÃ½ mÄ›sÃ­c.
    """
    summary = {}
    total_hours = 0
    uvazek_config = global_settings.get(uvazek, [])
    for entry in uvazek_config:
        shift_name = entry[0]
        summary[shift_name] = 0
    for i in range(1, len(day_plan_list)):
        shift = str(day_plan_list[i]).strip() if day_plan_list[i] is not None else ""
        for entry in uvazek_config:
            if shift == entry[0]:
                try:
                    hours = int(entry[1])
                except ValueError:
                    hours = 0
                total_hours += hours
                summary[shift] += 1
                break
    return total_hours, summary

# ----- NOVÃ‰: Funkce pro zamykÃ¡nÃ­ a odemykÃ¡nÃ­ plÃ¡nu
def lock_plan():
    """
    Zamkne plÃ¡n pro zadanÃ½ rok, mÄ›sÃ­c a smÄ›nu.
    """
    rok = combo_rok_smena.get().strip()
    mesic_display = combo_mesic_smena.get().strip()
    smena_val = combo_smena_smena.get().strip()
    if not (rok and mesic_display and smena_val):
        messagebox.showwarning("UpozornÄ›nÃ­", "Vyberte Rok, mÄ›sÃ­c a SmÄ›nu pro zamknutÃ­.")
        return
    mesic_map = {"Leden": "leden", "Ãšnor": "unor", "BÅ™ezen": "brezen", "Duben": "duben", "KvÄ›ten": "kveten",
                 "ÄŒerven": "cerven", "ÄŒervenec": "cervenec", "Srpen": "srpen", "ZÃ¡Å™Ã­": "zari",
                 "Å˜Ã­jen": "rijen", "Listopad": "listopad", "Prosinec": "prosinec"}
    internal_mesic = mesic_map.get(mesic_display, mesic_display.lower())
    key = f"{rok}_{smena_val}_{internal_mesic}"
    global_settings["locked_plans"][key] = True
    save_config(global_settings)
    messagebox.showinfo("ZÃ¡mek", f"PlÃ¡n {key} byl zamknut.")
    zobraz_plan_smeny()  # ObnovÃ­ zobrazenÃ­ plÃ¡nu

def unlock_plan():
    """
    Odemkne plÃ¡n pro zadanÃ½ rok, mÄ›sÃ­c a smÄ›nu.
    """
    rok = combo_rok_smena.get().strip()
    mesic_display = combo_mesic_smena.get().strip()
    smena_val = combo_smena_smena.get().strip()
    if not (rok and mesic_display and smena_val):
        messagebox.showwarning("UpozornÄ›nÃ­", "Vyberte Rok, MÄ›sÃ­c a SmÄ›nu pro odemknutÃ­.")
        return
    mesic_map = {"Leden": "leden", "Ãšnor": "unor", "BÅ™ezen": "brezen", "Duben": "duben", "KvÄ›ten": "kveten",
                 "ÄŒerven": "cerven", "ÄŒervenec": "cervenec", "Srpen": "srpen", "ZÃ¡Å™Ã­": "zari",
                 "Å˜Ã­jen": "rijen", "Listopad": "listopad", "Prosinec": "prosinec"}
    internal_mesic = mesic_map.get(mesic_display, mesic_display.lower())
    key = f"{rok}_{smena_val}_{internal_mesic}"
    if key in global_settings.get("locked_plans", {}):
        del global_settings["locked_plans"][key]
        save_config(global_settings)
        messagebox.showinfo("ZÃ¡mek", f"PlÃ¡n {key} byl odemknut.")
    else:
        messagebox.showinfo("ZÃ¡mek", "PlÃ¡n nenÃ­ zamknut.")
    zobraz_plan_smeny()  # ObnovÃ­ zobrazenÃ­ plÃ¡nu

# ----- Funkce pro zobrazenÃ­ dialogu s vÃ½bÄ›rem smÄ›ny
def ask_shift(allowed_shifts, current_value):
    """
    ZobrazÃ­ dialogovÃ© okno pro vÃ½bÄ›r smÄ›ny a vrÃ¡tÃ­ vybranou hodnotu.
    """
    dialog = tk.Toplevel()
    dialog.title("Vyberte smÄ›nu")
    tk.Label(dialog, text="Vyberte smÄ›nu:").pack(padx=10, pady=10)
    combo = ttk.Combobox(dialog, state="readonly", values=allowed_shifts)
    if current_value in allowed_shifts:
        combo.set(current_value)
    else:
        combo.set(allowed_shifts[0])
    combo.pack(padx=10, pady=5)
    result = {"value": None}
    def on_ok():
        result["value"] = combo.get()
        dialog.destroy()
    def on_cancel():
        dialog.destroy()
    btn_frame = tk.Frame(dialog)
    btn_frame.pack(padx=10, pady=10)
    tk.Button(btn_frame, text="OK", command=on_ok).pack(side=tk.LEFT, padx=5)
    tk.Button(btn_frame, text="ZruÅ¡it", command=on_cancel).pack(side=tk.LEFT, padx=5)
    dialog.grab_set()
    dialog.wait_window()
    return result["value"]

# ----- UpravenÃ¡ funkce pro vykreslenÃ­ mÄ›sÃ­ÄnÃ­ho plÃ¡nu s kontrolou zÃ¡mku
def render_month_grid(parent, year, month_num, plan_json, month_label, holidays, uvazek, smena, editable=False, highlight=False):
    """
    VykreslÃ­ mÅ™Ã­Å¾ku mÄ›sÃ­ÄnÃ­ho plÃ¡nu.
    Zahrnuje zÃ¡hlavÃ­ s dny v tÃ½dnu, ÄÃ­sla dnÅ¯ a hodnoty plÃ¡nu.
    PÅ™idÃ¡na kontrola, zda je plÃ¡n zamknutÃ½.
    """
    try:
        if highlight:
            frame = tk.LabelFrame(parent, text=month_label, font=("TkDefaultFont", 10, "bold"), bd=3, relief="groove")
        else:
            frame = tk.LabelFrame(parent, text=month_label, font=("TkDefaultFont", 10, "bold"))
        frame.pack(fill=tk.X, padx=5, pady=5)
        _, num_days = calendar.monthrange(int(year), month_num)
        try:
            day_plan_list = json.loads(plan_json)
        except Exception:
            day_plan_list = [""] * 32
        frame.day_plan_list = day_plan_list

        # MOD: Kontrola, zda je plÃ¡n zamknutÃ½
        mesic_map = {"Leden": "leden", "Ãšnor": "unor", "BÅ™ezen": "brezen", "Duben": "duben", "KvÄ›ten": "kveten",
                     "ÄŒerven": "cerven", "ÄŒervenec": "cervenec", "Srpen": "srpen", "ZÃ¡Å™Ã­": "zari",
                     "Å˜Ã­jen": "rijen", "Listopad": "listopad", "Prosinec": "prosinec"}
        internal_month = mesic_map.get(month_label, month_label.lower())
        key = f"{year}_{smena}_{internal_month}"
        if global_settings.get("locked_plans", {}).get(key, False):
            frame.is_locked = True
            tk.Label(frame, text="PlÃ¡n je zamknut ğŸ”’", font=("TkDefaultFont", 10, "bold"), fg="red")\
              .grid(row=4, column=0, columnspan=num_days, pady=5)
        else:
            frame.is_locked = False

        weekdays = ["Po", "Ãšt", "St", "ÄŒt", "PÃ¡", "So", "Ne"]

        # ----- VykreslenÃ­ zÃ¡hlavÃ­ s dny v tÃ½dnu
        for day in range(1, num_days + 1):
            current_date = date(int(year), month_num, day)
            weekday = current_date.weekday()
            if current_date in holidays:
                header_bg = "red"
                header_font = ("TkDefaultFont", 10, "bold")
            elif weekday >= 5:
                header_bg = "gray"
                header_font = ("TkDefaultFont", 10, "bold")
            else:
                header_bg = "white"
                header_font = ("TkDefaultFont", 10)
            abbrev = weekdays[weekday]
            tk.Label(frame, text=abbrev, borderwidth=1, relief="solid", width=4,
                     bg=header_bg, font=header_font).grid(row=0, column=day-1, padx=1, pady=1)
        # ----- VykreslenÃ­ ÄÃ­sel dnÅ¯
        for day in range(1, num_days + 1):
            current_date = date(int(year), month_num, day)
            weekday = current_date.weekday()
            if current_date in holidays:
                day_bg = "red"
            elif weekday >= 5:
                day_bg = "gray"
            else:
                day_bg = "white"
            tk.Label(frame, text=str(day), borderwidth=1, relief="solid", width=4,
                     bg=day_bg).grid(row=1, column=day-1, padx=1, pady=1)
        # ----- VykreslenÃ­ jednotlivÃ½ch hodnot plÃ¡nu
        for day in range(1, num_days + 1):
            current_date = date(int(year), month_num, day)
            weekday = calendar.weekday(int(year), month_num, day)
            bg = "white"
            if weekday >= 5:
                bg = "lightgray"
            if current_date in holidays:
                bg = "red"
            plan_value = ""
            if day < len(day_plan_list):
                plan_value = str(day_plan_list[day]) if day_plan_list[day] is not None else ""
            if plan_value:
                for entry in global_settings.get(uvazek, []):
                    if plan_value == entry[0]:
                        if entry[2].lower() != "neutrÃ¡lnÃ­":
                            bg = COLOR_MAP.get(entry[2], entry[2])
                        break
            if editable:
                widget = tk.Button(frame, text=plan_value, width=4, bg=bg)
                # MOD: pÅ™edÃ¡nÃ­ informace o zÃ¡mku do funkce editace buÅˆky
                widget.config(command=lambda w=widget, idx=day, dpl=frame.day_plan_list, uvazek=uvazek, locked=frame.is_locked:
                              edit_cell(w, idx, dpl, uvazek, locked))
            else:
                widget = tk.Label(frame, text=plan_value, borderwidth=1, relief="solid", width=4, bg=bg)
            widget.grid(row=2, column=day-1, padx=1, pady=1)
        # ----- VÃ½poÄet a vykreslenÃ­ souhrnu pro danÃ½ mÄ›sÃ­c
        total_hours, summary = calculate_month_summary(day_plan_list, uvazek)
        summary_text = f"Celkem: {total_hours} hodin | " + " | ".join([f"{shift}={count}" for shift, count in summary.items()])
        tk.Label(frame, text=summary_text, font=("TkDefaultFont", 10, "italic")).grid(row=3, column=0, columnspan=num_days, pady=5)
        return frame
    except Exception as e:
        logging.error(f"Chyba pÅ™i renderovÃ¡nÃ­ plÃ¡nu pro {month_label}: {e}")
        messagebox.showerror("Chyba", f"Chyba pÅ™i renderovÃ¡nÃ­ plÃ¡nu: {e}")

# ----- UpravenÃ¡ funkce pro Ãºpravu buÅˆky (plÃ¡n smÄ›ny)
def edit_cell(button, day_index, day_plan_list, uvazek, locked=False):
    """
    UmoÅ¾Åˆuje Ãºpravu hodnoty buÅˆky, pokud mÃ¡ uÅ¾ivatel dostateÄnÃ¡ oprÃ¡vnÄ›nÃ­.
    """
    if locked:
        messagebox.showwarning("UpozornÄ›nÃ­", "PlÃ¡n je zamknut ğŸ”’. Ãšpravy nejsou povoleny.")
        return
    old_value = button["text"].strip()  # Definice pÅ¯vodnÃ­ hodnoty
    if current_user_role in ["admin", "superadmin"]:
        pass  # Admin mÃ¡ plnÃ¡ prÃ¡va
    elif current_user_role == "velitel":
        if button["text"].strip() == "":
            messagebox.showerror("Chyba", "NemÃ¡te oprÃ¡vnÄ›nÃ­ mÄ›nit prÃ¡zdnou buÅˆku.")
            return
    else:
        messagebox.showerror("Chyba", "NemÃ¡te oprÃ¡vnÄ›nÃ­ k ÃºpravÄ›.")
        return
    allowed_shifts = [entry[0] for entry in global_settings.get(uvazek, [])]
    new_value = ask_shift(allowed_shifts, button["text"])
    if new_value is not None and new_value != old_value:
        # ZaznamenÃ¡nÃ­ zmÄ›ny do logu:
        # ZÃ­skÃ¡nÃ­ jmÃ©na zamÄ›stnance z comboboxu ZamÄ›stnanec
        employee_name = employee_combobox.get().strip() if employee_combobox.get() else "neznÃ¡mÃ½ zamÄ›stnanec"
        log_action(
            f"UÅ¾ivatel {current_user_name} ({current_user_role}) upravil plÃ¡n zamÄ›stnance {employee_name} "
            f"(Ãºvazek {uvazek}): den {day_index}, zmÄ›na z '{old_value}' na '{new_value}'"
        )
        button.config(text=new_value)
        day_plan_list[day_index] = new_value

# ----- Funkce pro aktualizaci seznamu zamÄ›stnancÅ¯
def update_employee_list():
    """
    NaÄte a aktualizuje seznam zamÄ›stnancÅ¯ z databÃ¡ze podle vybranÃ©ho roku.
    U velitele se naÄtou pouze zÃ¡znamy jeho smÄ›ny.
    """
    try:
        year_filter = year_combobox.get()
        query = "SELECT DISTINCT jmeno_prijmeni FROM plans WHERE 1=1"
        params = []
        if year_filter != "":
            query += " AND roky = ?"
            params.append(year_filter)
        # Pokud je aktuÃ¡lnÄ› pÅ™ihlÃ¡Å¡en velitel, naÄteme pouze jeho smÄ›nu
        if current_user_role == "velitel":
            shift_value = shift_filter_combobox.get()
            if shift_value != "":
                query += " AND smena = ?"
                params.append(shift_value)
        else:
            shift_filter = shift_filter_combobox.get()
            if shift_filter != "":
                query += " AND smena = ?"
                params.append(shift_filter)
        with sqlite3.connect("service_plans.db") as conn:
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            cursor.execute(query, params)
            employee_combobox['values'] = [row["jmeno_prijmeni"] for row in cursor.fetchall()]
    except Exception as e:
        logging.error(f"Chyba pÅ™i aktualizaci seznamu zamÄ›stnancÅ¯: {e}")
        messagebox.showerror("Chyba", f"Chyba pÅ™i aktualizaci seznamu zamÄ›stnancÅ¯: {e}")

# ----- VytvoÅ™enÃ­ sekce nastavenÃ­ pro danÃ½ Ãºvazek
def create_setting_section(parent, uvazek):
    """
    VytvoÅ™Ã­ GUI sekci pro nastavenÃ­ Ãºvazku, kde lze pÅ™idÃ¡vat Äi mazat smÄ›ny s pÅ™idruÅ¾enÃ½m poÄtem hodin a barvou.
    """
    section_frame = tk.LabelFrame(parent, text=f"Ãšvazek {uvazek}", font=("TkDefaultFont", 10, "bold"), width=180, height=220)
    section_frame.pack(side=tk.LEFT, padx=10, pady=5)
    section_frame.pack_propagate(False)

    listbox = tk.Listbox(section_frame, height=4)
    listbox.pack(fill=tk.X, padx=5, pady=5)

        # GlobÃ¡lnÃ­ binding: bez kontroly polohy, scrollujeme listbox vÅ¾dy
    section_frame.winfo_toplevel().bind_all("<MouseWheel>",
        lambda event: listbox.yview_scroll(-1 * (event.delta // 120), "units"))
     
    default_item = [" ", "0", "neutrÃ¡lnÃ­"]
    if default_item not in global_settings.get(uvazek, []):
        global_settings[uvazek].insert(0, default_item)
    for item in global_settings.get(uvazek, []):
        listbox.insert(tk.END, f"{item[0]} - {item[1]} hodin ({item[2]})")
    entry_frame = tk.Frame(section_frame)
    entry_frame.pack(fill=tk.X, padx=5, pady=5)
    tk.Label(entry_frame, text="SmÄ›na:").grid(row=0, column=0, sticky="e", padx=5, pady=2)
    shift_entry = tk.Entry(entry_frame, width=10)
    shift_entry.grid(row=0, column=1, padx=5, pady=2)
    tk.Label(entry_frame, text="Hodiny:").grid(row=1, column=0, sticky="e", padx=5, pady=2)
    hours_entry = tk.Entry(entry_frame, width=10)
    hours_entry.grid(row=1, column=1, padx=5, pady=2)
    tk.Label(entry_frame, text="Barva:").grid(row=2, column=0, sticky="e", padx=5, pady=2)
    colors = [
        "Å¾lutÃ©", "zelenÃ©", "oranÅ¾ovÃ©", "rÅ¯Å¾ovÃ©", "modrÃ©",
        "hnÄ›dÃ©", "neutrÃ¡lnÃ­", "fialovÃ©", "tyrkysovÃ©", "Å¡edÃ©"
    ]
    color_combo = ttk.Combobox(entry_frame, state="readonly", values=colors, width=10)
    color_combo.set("neutrÃ¡lnÃ­")
    color_combo.grid(row=2, column=1, padx=5, pady=2)
    def add_entry():
        shift = shift_entry.get().strip()
        hours = hours_entry.get().strip()
        color = color_combo.get().strip()
        if shift and hours and color:
            listbox.insert(tk.END, f"{shift} - {hours} hodin ({color})")
            global_settings[uvazek].append([shift, hours, color])
            shift_entry.delete(0, tk.END)
            hours_entry.delete(0, tk.END)
            color_combo.set("neutrÃ¡lnÃ­")
        else:
            messagebox.showwarning("UpozornÄ›nÃ­", "VyplÅˆte vÅ¡echny hodnoty.")
    def delete_entry():
        selection = listbox.curselection()
        if selection:
            index = selection[0]
            listbox.delete(index)
            global_settings[uvazek].pop(index)
        else:
            messagebox.showwarning("UpozornÄ›nÃ­", "Nevybrali jste zÃ¡znam k odstranÄ›nÃ­.")
    btn_frame = tk.Frame(section_frame)
    btn_frame.pack(fill=tk.X, padx=5, pady=5)
    tk.Button(btn_frame, text="PÅ™idat", command=add_entry).pack(side=tk.LEFT, padx=5)
    tk.Button(btn_frame, text="Smazat", command=delete_entry).pack(side=tk.LEFT, padx=5)
    return listbox

# ----- NaÄtenÃ­ seznamu zamÄ›stnancÅ¯ a rokÅ¯ z databÃ¡ze
def populate_employee_and_year():
    """
    NaÄte z databÃ¡ze seznam zamÄ›stnancÅ¯ a dostupnÃ½ch rokÅ¯, a aktualizuje pÅ™Ã­sluÅ¡nÃ¡ GUI pole.
    """
    try:
        with sqlite3.connect("service_plans.db") as conn:
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            cursor.execute("SELECT DISTINCT jmeno_prijmeni FROM plans")
            employee_combobox['values'] = [row["jmeno_prijmeni"] for row in cursor.fetchall()]
            cursor.execute("SELECT DISTINCT roky FROM plans")
            roky = [row["roky"] for row in cursor.fetchall()]
            current_year_str = str(datetime.now().year)
            if current_year_str in roky:
                year_combobox.set(current_year_str)
            else:
                year_combobox.set(current_year_str)
            cursor.execute("SELECT DISTINCT jmeno_prijmeni FROM plans")
            filter_jmeno['values'] = [row["jmeno_prijmeni"] for row in cursor.fetchall()]
            cursor.execute("SELECT DISTINCT roky FROM plans")
            filter_rok['values'] = [row["roky"] for row in cursor.fetchall()]
            cursor.execute("SELECT DISTINCT smena FROM plans")
            filter_smena['values'] = [row["smena"] for row in cursor.fetchall()]
    except Exception as e:
        logging.error(f"Chyba pÅ™i naÄÃ­tÃ¡nÃ­ hodnot pro filtry: {e}")
        messagebox.showerror("Chyba", f"Chyba pÅ™i naÄÃ­tÃ¡nÃ­ hodnot pro filtry: {e}")

# ----- PÅ™ihlaÅ¡ovacÃ­ funkce
def login():
    global current_user_role, current_user_name, current_user_shift
    nm = login_name_entry.get().strip()
    pwd = login_pwd_entry.get().strip()
    if nm == "" and pwd == "":
        current_user_role = "uÅ¾ivatel"
        current_user_name = "uÅ¾ivatel"
        login_status_label.config(text="PÅ™ihlÃ¡Å¡en: uÅ¾ivatel (uÅ¾ivatel)")
        log_action("PÅ™ihlÃ¡Å¡enÃ­ jako uÅ¾ivatel (vÃ½chozÃ­)")
        apply_access_control()
        messagebox.showinfo("PÅ™ihlÃ¡Å¡enÃ­", "PÅ™ihlÃ¡Å¡enÃ­ probÄ›hlo jako uÅ¾ivatel.")
        login_pwd_entry.delete(0, tk.END)
        for widget in plan_display_frame.winfo_children():
            widget.destroy()
        update_employee_list()
        return

    try:
        for entry in global_settings.get("access", []):
            if entry[0] == nm and decode_password(entry[1]) == pwd:
                current_user_role = entry[2]
                current_user_name = nm
                login_status_label.config(text=f"PÅ™ihlÃ¡Å¡en: {nm} ({current_user_role})")
                log_action(f"PÅ™ihlÃ¡Å¡enÃ­ jako {nm} ({current_user_role})")
                current_user_shift = None
                if current_user_role == "velitel":
                    parts = current_user_name.split()
                    if len(parts) >= 3 and parts[0] == "Velitel" and parts[1].lower() == "smÄ›ny":
                        shift_value = "SmÄ›na " + parts[2]
                        shift_filter_combobox.set(shift_value)
                        shift_filter_combobox.config(state="disabled")
                    else:
                        shift_filter_combobox.set("")
                    employee_combobox.set("")
                else:
                    shift_filter_combobox.config(state="readonly")
                apply_access_control()
                update_employee_list()  # Aktualizace seznamu zamÄ›stnancÅ¯
                for widget in plan_display_frame.winfo_children():
                    widget.destroy()
                messagebox.showinfo("PÅ™ihlÃ¡Å¡enÃ­", f"PÅ™ihlÃ¡Å¡enÃ­ probÄ›hlo ÃºspÄ›Å¡nÄ› jako {nm} ({current_user_role})")
                login_pwd_entry.delete(0, tk.END)
                return
        messagebox.showerror("PÅ™ihlÃ¡Å¡enÃ­", "NeplatnÃ© pÅ™ihlaÅ¡ovacÃ­ Ãºdaje")
    except Exception as e:
        logging.error(f"Chyba pÅ™i pÅ™ihlaÅ¡ovÃ¡nÃ­: {e}")
        messagebox.showerror("Chyba", f"Chyba pÅ™i pÅ™ihlaÅ¡ovÃ¡nÃ­: {e}")

# ----- Funkce pro odhlÃ¡Å¡enÃ­ uÅ¾ivatele
def logout():
    """
    Resetuje Ãºdaje o pÅ™ihlÃ¡Å¡enÃ©m uÅ¾ivateli a obnovÃ­ vÃ½chozÃ­ nastavenÃ­ GUI.
    """
    global current_user_role, current_user_name, current_user_shift
    current_user_role = "uÅ¾ivatel"
    current_user_name = "uÅ¾ivatel"
    current_user_shift = None
    login_status_label.config(text="Nejste pÅ™ihlÃ¡Å¡eni")
    login_name_entry.delete(0, tk.END)
    login_pwd_entry.delete(0, tk.END)
    shift_filter_combobox.config(state="readonly")
    shift_filter_combobox.set("")
    apply_access_control()
    update_employee_list()
    messagebox.showinfo("OdhlÃ¡Å¡enÃ­", "Byl jste ÃºspÄ›Å¡nÄ› odhlÃ¡Å¡en.")
    # ZavÅ™enÃ­ hlavnÃ­ho okna aplikace
    root.destroy()

# ----- Funkce pro nastavenÃ­ pÅ™Ã­stupovÃ½ch prÃ¡v podle role uÅ¾ivatele
def apply_access_control():
    """
    Skryje nebo zobrazÃ­ urÄitÃ© zÃ¡loÅ¾ky v GUI podle role pÅ™ihlÃ¡Å¡enÃ©ho uÅ¾ivatele.
    """
    try:
        role = current_user_role if current_user_role is not None else "uÅ¾ivatel"
        for i in range(notebook.index("end")):
            tab_text = notebook.tab(i, "text")
            if role in ["admin", "superadmin"]:
                notebook.tab(i, state="normal")
            else:
                if tab_text == "PlÃ¡ny":
                    notebook.tab(i, state="normal")
                else:
                    notebook.tab(i, state="hidden")
        for i in range(plans_notebook.index("end")):
            plans_notebook.tab(i, state="normal")
        if role in ["admin", "superadmin"]:
            # ZobrazÃ­ tlaÄÃ­tka, pokud jsou spravovÃ¡na metodou pack, nastavÃ­me pack() znovu
            btn_lock_plan.pack(side=tk.LEFT, padx=5)
            btn_unlock_plan.pack(side=tk.LEFT, padx=5)
        else:
            # Skryjeme tlaÄÃ­tka pomocÃ­ pack_forget()
            btn_lock_plan.pack_forget()
            btn_unlock_plan.pack_forget()
    except Exception as e:
        logging.error(f"Chyba pÅ™i nastavovÃ¡nÃ­ pÅ™Ã­stupovÃ½ch prÃ¡v: {e}")

        # MOD: Aktualizace tlaÄÃ­tek pro zÃ¡mek plÃ¡nu v zÃ¡loÅ¾ce SmÄ›na
        try:
            if current_user_role == "superadmin":
                btn_lock_plan.grid()  # zobrazÃ­ tlaÄÃ­tka
                btn_unlock_plan.grid()
            else:
                btn_lock_plan.grid_remove()
                btn_unlock_plan.grid_remove()
        except Exception as e:
            logging.error(f"Chyba pÅ™i aktualizaci tlaÄÃ­tek pro zÃ¡mek: {e}")
    except Exception as e:
        logging.error(f"Chyba pÅ™i nastavovÃ¡nÃ­ pÅ™Ã­stupovÃ½ch prÃ¡v: {e}")

# ----- OtevÅ™enÃ­ okna Fond hodin
def open_fond_window():
    fond_window = tk.Toplevel(root)
    fond_window.title("Fond hodin")
    fond_window.geometry("330x280")  # MÅ¯Å¾ete upravit velikost dle potÅ™eby
# --- VÃ½bÄ›r roku a smÄ›ny ---  
    tk.Label(fond_window, text="Rok:").grid(row=0, column=0, padx=5, pady=5, sticky="e")
    years = [str(y) for y in range(datetime.now().year - 5, datetime.now().year + 6)]
    year_combo = ttk.Combobox(fond_window, values=years, state="readonly", width=10)
    year_combo.grid(row=0, column=1, padx=5, pady=5)
    year_combo.set(str(datetime.now().year))
    
    tk.Label(fond_window, text="SmÄ›na:").grid(row=0, column=2, padx=5, pady=5, sticky="e")
    shifts = ["SmÄ›na 1", "SmÄ›na 2", "SmÄ›na 3", "SmÄ›na 4", "SmÄ›na 5", "SmÄ›na 6"]
    shift_combo = ttk.Combobox(fond_window, values=shifts, state="readonly", width=10)
    shift_combo.grid(row=0, column=3, padx=5, pady=5)
    shift_combo.set("SmÄ›na 1")
    
    # --- Pole pro mÄ›sÃ­ÄnÃ­ fondy (Leden aÅ¾ Prosinec) ---
    months = ["Leden", "Ãšnor", "BÅ™ezen", "Duben", "KvÄ›ten", "ÄŒerven",
              "ÄŒervenec", "Srpen", "ZÃ¡Å™Ã­", "Å˜Ã­jen", "Listopad", "Prosinec"]
    month_entries = {}
    for i, month in enumerate(months):
        row = i // 2 + 1
        col = (i % 2) * 2  # sloupec pro popisek a hodnotu
        tk.Label(fond_window, text=f"{month}:").grid(row=row, column=col, padx=5, pady=5, sticky="e")
        entry = tk.Entry(fond_window, width=10)
        entry.grid(row=row, column=col+1, padx=5, pady=5)
        month_entries[month.lower()] = entry  # klÃ­Ä v malÃ½ch pÃ­smenech
    
    # --- Funkce pro naÄtenÃ­ uloÅ¾enÃ½ch hodnot ---
    def load_fond_settings():
        selected_year = year_combo.get().strip()
        selected_shift = shift_combo.get().strip()
        if ("fond_hodin" in global_settings and 
            selected_year in global_settings["fond_hodin"] and 
            selected_shift in global_settings["fond_hodin"][selected_year]):
            fond_data = global_settings["fond_hodin"][selected_year][selected_shift]
            for month in months:
                key = month.lower()
                value = fond_data.get(key, "")
                month_entries[key].delete(0, tk.END)
                month_entries[key].insert(0, str(value))
        else:
            # Pokud pro vybranÃ½ rok a smÄ›nu jeÅ¡tÄ› nejsou uloÅ¾ena data, vymaÅ¾eme pole
            for entry in month_entries.values():
                entry.delete(0, tk.END)
    
    # --- Funkce pro uloÅ¾enÃ­ novÃ½ch hodnot ---
    def save_fond_settings():
        selected_year = year_combo.get().strip()
        selected_shift = shift_combo.get().strip()
        # Pokud klÃ­Ä jeÅ¡tÄ› neexistuje, vytvoÅ™Ã­me ho
        if "fond_hodin" not in global_settings:
            global_settings["fond_hodin"] = {}
        if selected_year not in global_settings["fond_hodin"]:
            global_settings["fond_hodin"][selected_year] = {}
        fond_data = {}
        for month in months:
            key = month.lower()
            try:
                val = int(month_entries[key].get().strip())
            except ValueError:
                val = 0
            fond_data[key] = val
        global_settings["fond_hodin"][selected_year][selected_shift] = fond_data
        save_config(global_settings)
        messagebox.showinfo("NastavenÃ­", "Fond hodin byl uloÅ¾en.")
    
    # --- TlaÄÃ­tka pro naÄtenÃ­ a uloÅ¾enÃ­ ---
    load_button = tk.Button(fond_window, text="NaÄÃ­st nastavenÃ­", command=load_fond_settings)
    load_button.grid(row=7, column=0, columnspan=2, padx=5, pady=10)
    save_button = tk.Button(fond_window, text="UloÅ¾it nastavenÃ­", command=save_fond_settings)
    save_button.grid(row=7, column=2, columnspan=2, padx=5, pady=10)

# ----- OtevÅ™enÃ­ okna globÃ¡lnÃ­ho nastavenÃ­
def open_settings_window():
    """
    ZobrazÃ­ okno pro globÃ¡lnÃ­ nastavenÃ­, kde lze mÄ›nit konfiguraci ÃºvazkÅ¯, pÅ™Ã­stupovÃ½ch prÃ¡v
    a takÃ© nastavit specifickÃ© roÄnÃ­ hodnoty pro smÄ›ny (I. a II. pololetÃ­).
    """
    try:
        settings_window = tk.Toplevel(root)
        settings_window.title("GlobÃ¡lnÃ­ nastavenÃ­")
        settings_window.geometry("700x730")
        settings_window.resizable(True, True)
        
        # --- Sekce nastavenÃ­ poÄÃ­tÃ¡nÃ­ hodin dle Ãºvazku (stÃ¡vajÃ­cÃ­)
        tk.Label(settings_window, text="NastavenÃ­ poÄÃ­tÃ¡nÃ­ hodin dle Ãšvazku", font=("TkDefaultFont", 12, "bold")).pack(pady=10)
        colors = [
            "Å¾lutÃ©", "zelenÃ©", "oranÅ¾ovÃ©", "rÅ¯Å¾ovÃ©", "modrÃ©",
            "hnÄ›dÃ©", "neutrÃ¡lnÃ­", "fialovÃ©", "tyrkysovÃ©", "Å¡edÃ©"
        ]
        uvazek_frame = tk.Frame(settings_window)
        uvazek_frame.pack(pady=5)
        listbox_40 = create_setting_section(uvazek_frame, "40")
        listbox_375 = create_setting_section(uvazek_frame, "37.5")
        listbox_3775 = create_setting_section(uvazek_frame, "37.75")
        
        # --- NovÃ¡ sekce pro nastavenÃ­ smÄ›n - VypuÅ¡tÄ›nÃ¡ a rannÃ­ smÄ›ny
        # Tato sekce umoÅ¾Åˆuje zadat pro zvolenÃ½ rok a smÄ›nu hodnoty pro I. a II. pololetÃ­.
        vypustena_frame = tk.LabelFrame(settings_window, text="NastavenÃ­ smÄ›n - VypuÅ¡tÄ›nÃ¡ smÄ›na a rannÃ­", font=("TkDefaultFont", 12, "bold"), width=380, height=280)
        vypustena_frame.pack(fill="both", expand=True, padx=10, pady=10)
        vypustena_frame.pack_propagate(False)
        
        # VÃ½bÄ›r roku
        tk.Label(vypustena_frame, text="Rok:", font=("TkDefaultFont", 10)).grid(row=0, column=0, padx=5, pady=5, sticky="e")
        vyber_rok = ttk.Combobox(vypustena_frame, state="readonly", width=10,
                                  values=[str(y) for y in range(datetime.now().year - 5, datetime.now().year + 6)])
        vyber_rok.grid(row=0, column=1, padx=5, pady=5, sticky="w")
        vyber_rok.set(str(datetime.now().year))
        
        # VÃ½bÄ›r smÄ›ny
        tk.Label(vypustena_frame, text="SmÄ›na:", font=("TkDefaultFont", 10)).grid(row=0, column=2, padx=5, pady=5, sticky="e")
        vyber_smÄ›na = ttk.Combobox(vypustena_frame, state="readonly", width=10,
                                    values=["SmÄ›na 1", "SmÄ›na 2", "SmÄ›na 3", "SmÄ›na 4", "SmÄ›na 5", "SmÄ›na 6"])
        vyber_smÄ›na.grid(row=0, column=3, padx=5, pady=5, sticky="w")
        vyber_smÄ›na.set("SmÄ›na 1")
        
        # Kontejner pro dynamicky naÄtenÃ½ formulÃ¡Å™
        formular_frame = tk.Frame(vypustena_frame)
        formular_frame.grid(row=1, column=4, rowspan=4, padx=10, pady=5)
        
        def nacist_formular():
            # SmazÃ¡nÃ­ starÃ©ho formulÃ¡Å™e, pokud existuje
            for widget in formular_frame.winfo_children():
                widget.destroy()
            
            # ZÃ­skÃ¡nÃ­ vybranÃ©ho roku a smÄ›ny
            rok = vyber_rok.get().strip()
            smena = vyber_smÄ›na.get().strip()
            saved_data = None
            if "year_settings" in global_settings and rok in global_settings["year_settings"]:
                saved_data = global_settings["year_settings"][rok].get(smena, None)
            
            # VytvoÅ™enÃ­ formulÃ¡Å™e pro I. pololetÃ­
            tk.Label(formular_frame, text="I. pololetÃ­", font=("TkDefaultFont", 10, "bold")).grid(row=0, column=0, columnspan=2, padx=5, pady=5)
            tk.Label(formular_frame, text="VypuÅ¡tÄ›nÃ¡ smÄ›na:", font=("TkDefaultFont", 10)).grid(row=1, column=0, padx=5, pady=5, sticky="e")
            entry_vypustena1 = tk.Entry(formular_frame, width=5)
            entry_vypustena1.grid(row=1, column=1, padx=5, pady=5, sticky="w")
            
            tk.Label(formular_frame, text="RannÃ­ smÄ›ny:", font=("TkDefaultFont", 10)).grid(row=2, column=0, padx=5, pady=5, sticky="e")
            entry_ranni1 = tk.Entry(formular_frame, width=5)
            entry_ranni1.grid(row=2, column=1, padx=5, pady=5, sticky="w")
            
            # VytvoÅ™enÃ­ formulÃ¡Å™e pro II. pololetÃ­
            tk.Label(formular_frame, text="II. pololetÃ­", font=("TkDefaultFont", 10, "bold")).grid(row=0, column=2, columnspan=2, padx=5, pady=5)
            tk.Label(formular_frame, text="VypuÅ¡tÄ›nÃ¡ smÄ›na:", font=("TkDefaultFont", 10)).grid(row=1, column=2, padx=5, pady=5, sticky="e")
            entry_vypustena2 = tk.Entry(formular_frame, width=5)
            entry_vypustena2.grid(row=1, column=3, padx=5, pady=5, sticky="w")
            
            tk.Label(formular_frame, text="RannÃ­ smÄ›ny:", font=("TkDefaultFont", 10)).grid(row=2, column=2, padx=5, pady=5, sticky="e")
            entry_ranni2 = tk.Entry(formular_frame, width=5)
            entry_ranni2.grid(row=2, column=3, padx=5, pady=5, sticky="w")
            
            # Pokud existujÃ­ uloÅ¾enÃ¡ data, pÅ™edvyplnÃ­me je
            if saved_data:
                entry_vypustena1.insert(0, str(saved_data.get("pololeti1", {}).get("vypustena", "")))
                entry_ranni1.insert(0, str(saved_data.get("pololeti1", {}).get("ranni", "")))
                entry_vypustena2.insert(0, str(saved_data.get("pololeti2", {}).get("vypustena", "")))
                entry_ranni2.insert(0, str(saved_data.get("pololeti2", {}).get("ranni", "")))
            
            # UloÅ¾enÃ­ odkazÅ¯ na widgety pro pozdÄ›jÅ¡Ã­ naÄtenÃ­ hodnot
            formular_frame.entries = {
                "vypustena1": entry_vypustena1,
                "ranni1": entry_ranni1,
                "vypustena2": entry_vypustena2,
                "ranni2": entry_ranni2
            }
        
        # TlaÄÃ­tko pro naÄtenÃ­ formulÃ¡Å™e dle vybranÃ½ch hodnot
        btn_nacist = tk.Button(vypustena_frame, text="NaÄÃ­st formulÃ¡Å™", command=nacist_formular)
        btn_nacist.grid(row=0, column=4, padx=5, pady=5)
        
        # TlaÄÃ­tko pro uloÅ¾enÃ­ nastavenÃ­ z formulÃ¡Å™e
        def ulozit_nastaveni_smÄ›ny():
            rok = vyber_rok.get().strip()
            smena = vyber_smÄ›na.get().strip()
            if not rok or not smena:
                messagebox.showerror("Chyba", "Vyberte rok a smÄ›nu.")
                return
                # PÅ™idÃ¡na kontrola, zda uÅ¾ byl naÄten formulÃ¡Å™
            if not hasattr(formular_frame, "entries"):
                messagebox.showerror("Chyba", "Nejprve naÄtÄ›te formulÃ¡Å™ kliknutÃ­m na 'NaÄÃ­st formulÃ¡Å™'.")
                return
            entries = formular_frame.entries
            data = {
                "vypustena1": int(entries["vypustena1"].get()) if entries["vypustena1"].get().isdigit() else 0,
                "ranni1": int(entries["ranni1"].get()) if entries["ranni1"].get().isdigit() else 0,
                "vypustena2": int(entries["vypustena2"].get()) if entries["vypustena2"].get().isdigit() else 0,
                "ranni2": int(entries["ranni2"].get()) if entries["ranni2"].get().isdigit() else 0,
            }
            if "year_settings" not in global_settings:
                global_settings["year_settings"] = {}
            if rok not in global_settings["year_settings"]:
                global_settings["year_settings"][rok] = {}
            global_settings["year_settings"][rok][smena] = {
                        "pololeti1": {
                        "vypustena": data["vypustena1"],
                        "ranni": data["ranni1"]
                    },
                    "pololeti2": {
                        "vypustena": data["vypustena2"],
                        "ranni": data["ranni2"]
                    }
            }
            save_config(global_settings)
            messagebox.showinfo("NastavenÃ­", "NastavenÃ­ smÄ›ny bylo uloÅ¾eno.")
        
        btn_ulozit_sm = tk.Button(vypustena_frame, text="UloÅ¾it nastavenÃ­ smÄ›ny", command=ulozit_nastaveni_smÄ›ny)
        btn_ulozit_sm.grid(row=5, column=0, columnspan=5, pady=10)
        
        # --- Sekce nastavenÃ­ pÅ™Ã­stupu (stÃ¡vajÃ­cÃ­)
        access_frame = tk.LabelFrame(settings_window, text="PÅ™Ã­stup", font=("TkDefaultFont", 10, "bold"), width=800, height=250)
        access_frame.pack(pady=10)
        access_frame.pack_propagate(False)
        tk.Label(access_frame, text="Heslo:").grid(row=0, column=0, padx=5, pady=2, sticky="e")
        pwd_entry = tk.Entry(access_frame, width=12, show="*")
        pwd_entry.grid(row=0, column=1, padx=5, pady=2)
        tk.Label(access_frame, text="JmÃ©no:").grid(row=1, column=0, padx=5, pady=2, sticky="e")
        name_entry = tk.Entry(access_frame, width=12)
        name_entry.grid(row=1, column=1, padx=5, pady=2)
        tk.Label(access_frame, text="Role:").grid(row=2, column=0, padx=5, pady=2, sticky="e")
        role_combo = ttk.Combobox(access_frame, state="readonly", values=["uÅ¾ivatel", "velitel", "admin", "superadmin"], width=10)
        role_combo.set("uÅ¾ivatel")
        role_combo.grid(row=2, column=1, padx=5, pady=2)
        access_listbox = tk.Listbox(access_frame, height=3)
        access_listbox.grid(row=0, column=2, rowspan=3, padx=5, pady=2)
        for item in global_settings.get("access", []):
            access_listbox.insert(tk.END, f"{item[0]} - {item[1]} - {item[2]}")
        access_btn_frame = tk.Frame(access_frame)
        access_btn_frame.grid(row=3, column=0, columnspan=3, pady=5)
        def add_access():
            pwd = pwd_entry.get().strip()
            nm = name_entry.get().strip()
            role = role_combo.get().strip()
            if pwd and nm and role:
                encoded_pwd = encode_password(pwd)
                access_listbox.insert(tk.END, f"{nm} - {encoded_pwd} - {role}")
                if "access" not in global_settings:
                    global_settings["access"] = []
                global_settings["access"].append((nm, encoded_pwd, role))
                pwd_entry.delete(0, tk.END)
                name_entry.delete(0, tk.END)
                role_combo.set("uÅ¾ivatel")
            else:
                messagebox.showwarning("UpozornÄ›nÃ­", "VyplÅˆte Heslo, JmÃ©no a Role.")
        def delete_access():
            selection = access_listbox.curselection()
            if selection:
                index = selection[0]
                access_listbox.delete(index)
                global_settings["access"].pop(index)
            else:
                messagebox.showwarning("UpozornÄ›nÃ­", "Nevybrali jste zÃ¡znam k odstranÄ›nÃ­.")
        tk.Button(access_btn_frame, text="PÅ™idat", command=add_access).pack(side=tk.LEFT, padx=5)
        tk.Button(access_btn_frame, text="Smazat", command=delete_access).pack(side=tk.LEFT, padx=5)

        def save_global_settings():
            save_config(global_settings)
            messagebox.showinfo("NastavenÃ­", "UloÅ¾enÃ­ nastavenÃ­ probÄ›hlo ÃºspÄ›Å¡nÄ›")
            settings_window.destroy()
        tk.Button(settings_window, text="UloÅ¾it nastavenÃ­", command=save_global_settings).pack(pady=20, anchor="center")
    except Exception as e:
        logging.error(f"Chyba v oknÄ› globÃ¡lnÃ­ho nastavenÃ­: {e}")
        messagebox.showerror("Chyba", f"Chyba v oknÄ› nastavenÃ­: {e}")

# ----- Funkce pro naÄtenÃ­ novÃ©ho plÃ¡nu z Excelu a jeho uloÅ¾enÃ­ do databÃ¡ze
def on_new():
    """
    NaÄte ExcelovÃ½ soubor, extrahuje z nÄ›j data plÃ¡nu a uloÅ¾Ã­ je do databÃ¡ze.
    """
    try:
        log_action("Stisknuto tlaÄÃ­tko 'NovÃ½'")
        file_path = filedialog.askopenfilename(title="Vyberte Excel soubor", filetypes=[("Excel soubory", "*.xlsx;*.xls")])
        if file_path:
            try:
                wb = load_workbook(filename=file_path, data_only=True)
                ws = wb.active
                osobni_cislo = ws["A2"].value
                jmeno_prijmeni = ws["C2"].value
                smena = "SmÄ›na "   # bude doplnÄ›no pÅ™i editaci
                uvazek = "37.5"    # pÅ™edpoklÃ¡dÃ¡me, Å¾e Ãºvazek je "40"
                cell_b30 = ws["B30"].value
                if isinstance(cell_b30, str):
                    if len(cell_b30) >= 4:
                        roky = cell_b30[-4:]
                    else:
                        messagebox.showerror("Chyba", "BuÅˆka B30 obsahuje Å™etÄ›zec, ale nenÃ­ dostateÄnÄ› dlouhÃ½ pro rok.")
                        return
                elif isinstance(cell_b30, (datetime, date)):
                    roky = str(cell_b30.year)
                else:
                    messagebox.showerror("Chyba", "BuÅˆka B30 neobsahuje platnÃ½ Ãºdaj o roce.")
                    return
                poradi = "0"
                month_rows = {
                    "leden": 4,
                    "unor": 6,
                    "brezen": 8,
                    "duben": 10,
                    "kveten": 12,
                    "cerven": 14,
                    "cervenec": 16,
                    "srpen": 18,
                    "zari": 20,
                    "rijen": 22,
                    "listopad": 24,
                    "prosinec": 26
                }
                month_data = load_month_data(ws, month_rows)
                with sqlite3.connect("service_plans.db") as conn:
                    conn.row_factory = sqlite3.Row
                    cursor = conn.cursor()
                    cursor.execute(
                        "SELECT id FROM plans WHERE jmeno_prijmeni = ? AND osobni_cislo = ? AND roky = ?",
                        (jmeno_prijmeni, osobni_cislo, roky)
                    )
                    existing_record = cursor.fetchone()
                    if existing_record is not None:
                        if not messagebox.askyesno("PotvrzenÃ­ pÅ™epsÃ¡nÃ­",
                                                   f"PlÃ¡n pro {jmeno_prijmeni} ({osobni_cislo}) pro rok {roky} jiÅ¾ existuje.\nChcete jej pÅ™epsat?"):
                            return
                        update_query = """
                            UPDATE plans
                            SET smena = ?, uvazek = ?, poradi = ?,
                                stara_dovolena = "0", dovolena = "0",
                                leden = ?, unor = ?, brezen = ?, duben = ?,
                                kveten = ?, cerven = ?, cervenec = ?, srpen = ?,
                                zari = ?, rijen = ?, listopad = ?, prosinec = ?
                            WHERE id = ?
                        """
                        params = (
                            smena, uvazek, poradi,
                            month_data["leden"], month_data["unor"], month_data["brezen"],
                            month_data["duben"], month_data["kveten"], month_data["cerven"],
                            month_data["cervenec"], month_data["srpen"], month_data["zari"],
                            month_data["rijen"], month_data["listopad"], month_data["prosinec"],
                            existing_record["id"]
                        )
                        cursor.execute(update_query, params)
                        conn.commit()
                        log_action("PlÃ¡n byl pÅ™epsÃ¡n v databÃ¡zi")
                        messagebox.showinfo("ÃšspÄ›ch", "PlÃ¡n byl ÃºspÄ›Å¡nÄ› pÅ™epsÃ¡n.")
                        refresh_treeview()
                        populate_employee_and_year()
                        return
                    else:
                        cursor.execute("""
                            INSERT INTO plans (
                                osobni_cislo, jmeno_prijmeni, smena, uvazek, roky, poradi,
                                stara_dovolena, dovolena,
                                leden, unor, brezen, duben, kveten, cerven, cervenec, srpen, zari, rijen, listopad, prosinec
                            ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                        """, (
                            osobni_cislo, jmeno_prijmeni, smena, uvazek, roky, poradi,
                            "0", "0",
                            month_data["leden"], month_data["unor"], month_data["brezen"],
                            month_data["duben"], month_data["kveten"], month_data["cerven"],
                            month_data["cervenec"], month_data["srpen"], month_data["zari"],
                            month_data["rijen"], month_data["listopad"], month_data["prosinec"]
                        ))
                        conn.commit()
                log_action("Data byla ÃºspÄ›Å¡nÄ› uloÅ¾ena do databÃ¡ze")
                messagebox.showinfo("ÃšspÄ›ch", "Data byla ÃºspÄ›Å¡nÄ› naÄtena a uloÅ¾ena do databÃ¡ze.")
                refresh_treeview()
                populate_employee_and_year()
            except Exception as e:
                logging.error(f"Chyba pÅ™i naÄÃ­tÃ¡nÃ­ dat z Excelu: {e}")
                messagebox.showerror("Chyba", f"DoÅ¡lo k chybÄ› pÅ™i naÄÃ­tÃ¡nÃ­ dat: {e}")
        else:
            print("Nebyl vybrÃ¡n Å¾Ã¡dnÃ½ soubor.")
    except Exception as e:
        logging.error(f"ObecnÃ¡ chyba pÅ™i volÃ¡nÃ­ on_new: {e}")
        messagebox.showerror("Chyba", f"DoÅ¡lo k chybÄ›: {e}")

# ----- Funkce pro Ãºpravu existujÃ­cÃ­ho zÃ¡znamu (Edit)
def on_edit():
    """
    UmoÅ¾Åˆuje upravit celÃ½ zÃ¡znam, coÅ¾ mohou provÃ¡dÄ›t pouze admin a superadmin.
    """
    try:
        if current_user_role not in ["admin", "superadmin"]:
            messagebox.showerror("Chyba", "Pouze admin a superadmin mohou mÄ›nit celÃ½ zÃ¡znam.")
            return
        log_action("Stisknuto tlaÄÃ­tko 'Edit'")
        selected_items = tree.selection()
        if not selected_items:
            messagebox.showwarning("UpozornÄ›nÃ­", "Nevybrali jste Å¾Ã¡dnou poloÅ¾ku k ÃºpravÄ›.")
            return
        record_id = selected_items[0]
        with sqlite3.connect("service_plans.db") as conn:
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM plans WHERE id = ?", (record_id,))
            row = cursor.fetchone()
        if row is None:
            messagebox.showerror("Chyba", "ZÃ¡znam nebyl nalezen.")
            return
        detail_window = tk.Toplevel(root)
        detail_window.geometry("600x250")
        detail_window.title("Editace zÃ¡znamu")
        detail_window.resizable(True, True)
        top_frame = tk.Frame(detail_window)
        top_frame.pack(side=tk.TOP, fill=tk.X, padx=10, pady=10)
        basic_frame = tk.Frame(top_frame)
        basic_frame.pack(side=tk.LEFT, fill=tk.X, expand=True)
        basic_left = tk.Frame(basic_frame)
        basic_left.grid(row=0, column=0, sticky="nw", padx=5, pady=2)
        basic_right = tk.Frame(basic_frame)
        basic_right.grid(row=0, column=1, sticky="nw", padx=20, pady=2)
        tk.Label(basic_left, text="OsobnÃ­ ÄÃ­slo:", font=("TkDefaultFont", 10, "bold")).grid(row=0, column=0, sticky="w", padx=5, pady=2)
        entry_osobni = tk.Entry(basic_left)
        entry_osobni.insert(0, str(row["osobni_cislo"]))
        entry_osobni.grid(row=0, column=1, sticky="w", padx=5, pady=2)
        tk.Label(basic_right, text="StarÃ¡ dovolenÃ¡ hodiny:", font=("TkDefaultFont", 10, "bold")).grid(row=0, column=0, sticky="e", padx=5, pady=2)
        entry_stara = tk.Entry(basic_right)
        entry_stara.insert(0, str(row["stara_dovolena"]))
        entry_stara.grid(row=0, column=1, sticky="w", padx=5, pady=2)
        tk.Label(basic_right, text="DovolenÃ¡ hodiny:", font=("TkDefaultFont", 10, "bold")).grid(row=1, column=0, sticky="e", padx=5, pady=2)
        entry_dovolena = tk.Entry(basic_right)
        entry_dovolena.insert(0, str(row["dovolena"]))
        entry_dovolena.grid(row=1, column=1, sticky="w", padx=5, pady=2)
        next_row = 1
        remaining_fields = {
            "JmÃ©no a pÅ™Ã­jmenÃ­": row["jmeno_prijmeni"],
            "SmÄ›na": row["smena"],
            "Ãšvazek": row["uvazek"],
            "Roky": row["roky"],
            "PoÅ™adÃ­ ve smÄ›nÄ›": row["poradi"]
        }
        basic_entries = {"OsobnÃ­ ÄÃ­slo": entry_osobni,
                         "StarÃ¡ dovolenÃ¡ hodiny": entry_stara,
                         "DovolenÃ¡ hodiny": entry_dovolena}
        for label_text, value in remaining_fields.items():
            tk.Label(basic_left, text=f"{label_text}:", font=("TkDefaultFont", 10, "bold")).grid(row=next_row, column=0, sticky="w", padx=5, pady=2)
            if label_text in ["SmÄ›na", "Ãšvazek"]:
                combo = ttk.Combobox(basic_left, state="normal")
                if label_text == "SmÄ›na":
                    combo['values'] = ["SmÄ›na 1", "SmÄ›na 2", "SmÄ›na 3", "SmÄ›na 4", "SmÄ›na 5", "SmÄ›na 6"]
                else:
                    combo['values'] = ["40", "37.5", "37.75"]
                combo.set(value)
                combo.grid(row=next_row, column=1, sticky="w", padx=5, pady=2)
                basic_entries[label_text] = combo
            elif label_text == "PoÅ™adÃ­ ve smÄ›nÄ›":
                combo = ttk.Combobox(basic_left, state="normal", values=[str(x) for x in range(0, 21)])
                combo.set(value)
                combo.grid(row=next_row, column=1, sticky="w", padx=5, pady=2)
                basic_entries[label_text] = combo
            else:
                entry = tk.Entry(basic_left)
                entry.insert(0, str(value))
                entry.grid(row=next_row, column=1, sticky="w", padx=5, pady=2)
                basic_entries[label_text] = entry
            next_row += 1
        bottom_frame = tk.Frame(detail_window)
        bottom_frame.pack(side=tk.BOTTOM, fill=tk.X, pady=10)
        save_button = tk.Button(bottom_frame, text="UloÅ¾it zmÄ›ny", 
                                command=lambda: save_changes(basic_entries, row, detail_window, record_id))
        save_button.pack(anchor="center")
        def save_changes(basic_entries, row, detail_window, record_id):
            try:
                updated_poradi = basic_entries["PoÅ™adÃ­ ve smÄ›nÄ›"].get()
                updated_osobni = basic_entries["OsobnÃ­ ÄÃ­slo"].get()
                updated_stara = basic_entries["StarÃ¡ dovolenÃ¡ hodiny"].get()
                updated_dovolena = basic_entries["DovolenÃ¡ hodiny"].get()
                updated_jmeno = basic_entries["JmÃ©no a pÅ™Ã­jmenÃ­"].get()
                updated_smena = basic_entries["SmÄ›na"].get()
                updated_uvazek = basic_entries["Ãšvazek"].get()
                updated_roky = basic_entries["Roky"].get()
                with sqlite3.connect("service_plans.db") as conn:
                    conn.row_factory = sqlite3.Row
                    cursor = conn.cursor()
                    update_query = """
                        UPDATE plans
                        SET osobni_cislo = ?, stara_dovolena = ?, dovolena = ?, jmeno_prijmeni = ?, smena = ?, uvazek = ?, roky = ?, poradi = ?
                        WHERE id = ?
                    """
                    params = [
                        updated_osobni, updated_stara, updated_dovolena, updated_jmeno,
                        updated_smena, updated_uvazek, updated_roky, updated_poradi,
                        record_id
                    ]
                    cursor.execute(update_query, params)
                    conn.commit()
                log_action(f"PlÃ¡n byl upraven uÅ¾ivatelem {current_user_name}")
                messagebox.showinfo("ÃšspÄ›ch", "ZmÄ›ny byly ÃºspÄ›Å¡nÄ› uloÅ¾eny.")
                detail_window.destroy()
                refresh_treeview()
            except Exception as e:
                logging.error(f"Chyba pÅ™i uklÃ¡dÃ¡nÃ­ zmÄ›n v editaci zÃ¡znamu: {e}")
                messagebox.showerror("Chyba", f"DoÅ¡lo k chybÄ› pÅ™i uklÃ¡dÃ¡nÃ­ zmÄ›n: {e}")
    except Exception as e:
        logging.error(f"Chyba pÅ™i naÄÃ­tÃ¡nÃ­ detailu zÃ¡znamu: {e}")
        messagebox.showerror("Chyba", f"DoÅ¡lo k chybÄ› pÅ™i naÄÃ­tÃ¡nÃ­ detailu: {e}")

# ----- Funkce pro smazÃ¡nÃ­ zÃ¡znamu
def on_delete():
    """
    SmaÅ¾e vybranÃ½ zÃ¡znam ze seznamu a databÃ¡ze.
    """
    try:
        log_action("Stisknuto tlaÄÃ­tko 'Smazat'")
        selected_items = tree.selection()
        if not selected_items:
            messagebox.showwarning("UpozornÄ›nÃ­", "Nevybrali jste Å¾Ã¡dnou poloÅ¾ku k smazÃ¡nÃ­.")
            return
        if not messagebox.askyesno("PotvrzenÃ­", "Opravdu chcete smazat vybranÃ©ho zamÄ›stnance?"):
            return
        for item in tree.selection():
            try:
                with sqlite3.connect("service_plans.db") as conn:
                    cursor = conn.cursor()
                    cursor.execute("DELETE FROM plans WHERE id = ?", (item,))
                    conn.commit()
                tree.delete(item)
                log_action(f"ZÃ¡znam s ID {item} byl smazÃ¡n z databÃ¡ze")
            except Exception as e:
                logging.error(f"Chyba pÅ™i mazÃ¡nÃ­ zÃ¡znamu s ID {item}: {e}")
                messagebox.showerror("Chyba", f"DoÅ¡lo k chybÄ› pÅ™i mazÃ¡nÃ­ zÃ¡znamu: {e}")
        messagebox.showinfo("ÃšspÄ›ch", "VybranÃ¡ poloÅ¾ka byla ÃºspÄ›Å¡nÄ› smazÃ¡na.")
    except Exception as e:
        logging.error(f"ObecnÃ¡ chyba pÅ™i mazÃ¡nÃ­: {e}")
        messagebox.showerror("Chyba", f"DoÅ¡lo k chybÄ›: {e}")

# ----- Funkce pro zobrazenÃ­ plÃ¡nu smÄ›ny v zÃ¡loÅ¾ce ZamÄ›stnanec
def show_employee_plan():
    """
    ZobrazÃ­ detailnÃ­ plÃ¡n smÄ›ny vybranÃ©ho zamÄ›stnance pro aktuÃ¡lnÃ­ rok.
    """
    try:
        global current_record_id, month_frames
        for widget in plan_display_frame.winfo_children():
            widget.destroy()
        selected_employee = employee_combobox.get()
        current_year = datetime.now().year
        year_combobox.set(str(current_year))
        selected_year = str(current_year)
        if not selected_employee or not selected_year:
            messagebox.showwarning("UpozornÄ›nÃ­", "Vyberte min. ZamÄ›stnance a Rok.")
            return
        with sqlite3.connect("service_plans.db") as conn:
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM plans WHERE jmeno_prijmeni = ? AND roky = ?", (selected_employee, selected_year))
            record = cursor.fetchone()
        if record is None:
            messagebox.showinfo("Informace", "Pro vybranÃ©ho ZamÄ›stnance a Rok nebyl nalezen Å¾Ã¡dnÃ½ plÃ¡n.")
            return

        jmeno = record["jmeno_prijmeni"]
        osobni = record["osobni_cislo"]
        smena = record["smena"]
        uvazek = record["uvazek"]
        nova_dovolena = record["dovolena"]
        stara_dovolena = record["stara_dovolena"]
        try:
            celkem_dovolena = float(nova_dovolena) + float(stara_dovolena)
        except Exception:
            celkem_dovolena = 0.0

        # Definice mÄ›sÃ­cÅ¯ pro prvnÃ­ a druhÃ© pololetÃ­
        half1_months = ["leden", "unor", "brezen", "duben", "kveten", "cerven"]
        half2_months = ["cervenec", "srpen", "zari", "rijen", "listopad", "prosinec"]

        # PoÄÃ­tÃ¡nÃ­ vÃ½skytÅ¯ znaku "-" pro Klouz I. a Klouz II.
        half1_dash_count = 0
        for m in half1_months:
            try:
                plan_list = json.loads(record[m])
            except Exception:
                plan_list = [""] * 32
            for cell in plan_list[1:]:
                if isinstance(cell, str):
                    half1_dash_count += cell.count("-")
        
        half2_dash_count = 0
        for m in half2_months:
            try:
                plan_list = json.loads(record[m])
            except Exception:
                plan_list = [""] * 32
            for cell in plan_list[1:]:
                if isinstance(cell, str):
                    half2_dash_count += cell.count("-")
        
        # PoÄÃ­tÃ¡nÃ­ vÃ½skytÅ¯ znaku "r" pro sloupce r I. a r II.
        half1_r_count = 0
        for m in half1_months:
            try:
                plan_list = json.loads(record[m])
            except Exception:
                plan_list = [""] * 32
            for cell in plan_list[1:]:
                if isinstance(cell, str):
                    half1_r_count += cell.count("r")
        
        half2_r_count = 0
        for m in half2_months:
            try:
                plan_list = json.loads(record[m])
            except Exception:
                plan_list = [""] * 32
            for cell in plan_list[1:]:
                if isinstance(cell, str):
                    half2_r_count += cell.count("r")

        # NaÄtenÃ­ globÃ¡lnÃ­ch hodnot z nastavenÃ­
        # Pro "Klouz I." a "Klouz II." se Äte pod klÃ­Äi "vypustena"
        # Pro "r I." a "r II." se Äte pod klÃ­Äi "ranni"
        klouz1_value = 0
        klouz2_value = 0
        ranni1_value = 0
        ranni2_value = 0
        year_str = str(selected_year)
        shift_name = record["smena"]
        if "year_settings" in global_settings:
            if year_str in global_settings["year_settings"]:
                if shift_name in global_settings["year_settings"][year_str]:
                    klouz1_value = global_settings["year_settings"][year_str][shift_name].get("pololeti1", {}).get("vypustena", 0)
                    klouz2_value = global_settings["year_settings"][year_str][shift_name].get("pololeti2", {}).get("vypustena", 0)
                    ranni1_value = global_settings["year_settings"][year_str][shift_name].get("pololeti1", {}).get("ranni", 0)
                    ranni2_value = global_settings["year_settings"][year_str][shift_name].get("pololeti2", {}).get("ranni", 0)
        
         # VytvoÅ™enÃ­ kompozitnÃ­ch Å™etÄ›zcÅ¯ pro sloupce
        klouz1_display = f"{klouz1_value}/{half1_dash_count}"
        klouz2_display = f"{klouz2_value}/{half2_dash_count}"
        ranni1_display = f"{ranni1_value}/{half1_r_count}"
        ranni2_display = f"{ranni2_value}/{half2_r_count}"

        # VÃ½poÄet plÃ¡novanÃ© dovolenÃ©
        dov_count = 0
        month_keys = ["leden", "unor", "brezen", "duben", "kveten", "cerven", 
                      "cervenec", "srpen", "zari", "rijen", "listopad", "prosinec"]
        uvazek_shifts = global_settings.get(uvazek, [])
        for key in month_keys:
            try:
                plan = json.loads(record[key])
            except Exception:
                plan = []
            for cell in plan[1:]:
                if isinstance(cell, str) and cell.strip() == "Dov":
                    dov_count += 1
        dov_hours = 0.0
        for entry in uvazek_shifts:
            if entry[0] == "Dov":
                try:
                    dov_hours = float(entry[1])
                except Exception:
                    dov_hours = 0.0
                break
        planned_vacation_hours = dov_count * dov_hours
        rozdil = celkem_dovolena - planned_vacation_hours

        # VÃ½poÄet celkovÃ©ho poÄtu smÄ›n dle uvazku:
        # Pro kaÅ¾dÃ½ den ve vÅ¡ech mÄ›sÃ­cÃ­ch se prochÃ¡zÃ­ symboly,
        # pÅ™iÄemÅ¾ pokud symbol nenÃ­ "Dov" a odpovÃ­dÃ¡ zÃ¡znamu v globÃ¡lnÃ­m nastavenÃ­ s hodinami > 1,
        # zapoÄÃ­tÃ¡ se tato instance.
        total_shifts = 0
        all_months = ["leden", "unor", "brezen", "duben", "kveten", "cerven",
                      "cervenec", "srpen", "zari", "rijen", "listopad", "prosinec"]
        for month in all_months:
            try:
                day_plan_list = json.loads(record[month])
            except Exception:
                day_plan_list = [""] * 32
            for cell in day_plan_list[1:]:
                if isinstance(cell, str):
                    symbol = cell.strip()
                    if symbol == "Dov":
                        continue
                    for entry in global_settings.get(uvazek, []):
                        try:
                            hours = float(entry[1])
                        except Exception:
                            hours = 0
                        if symbol == entry[0] and hours > 1:
                            total_shifts += 1
                            break
        celkem_smen_display = f"{total_shifts}"

        # SestavenÃ­ seznamu popiskÅ¯ a hodnot pro informaÄnÃ­ tabulku
        popisky = [
            "JmÃ©no a pÅ™Ã­jmenÃ­", "NovÃ¡ dovolenÃ¡", "StarÃ¡ dovolenÃ¡", "Celkem dovolenÃ¡", 
            "NaplÃ¡novat Dov", "RozdÃ­l PlÃ¡n a NÃ¡rok", "Celkem smÄ›n", 
            "Klouz I.", "Klouz II.", "r I.", "r II."
        ]
        hodnoty = [
            jmeno,
            f"{nova_dovolena} hod.",
            f"{stara_dovolena} hod.",
            f"{celkem_dovolena} hod.",
            f"{planned_vacation_hours} hod.",
            f"{rozdil} hod.",
            celkem_smen_display,
            klouz1_display,
            klouz2_display,
            ranni1_display,
            ranni2_display
        ]

        # VykreslenÃ­ informaÄnÃ­ tabulky
        info_table_frame = tk.Frame(plan_display_frame, bd=2, relief="groove")
        info_table_frame.pack(fill=tk.X, padx=10, pady=10)
        columns = len(popisky)
        for col in range(columns):
            lbl = tk.Label(info_table_frame, text=popisky[col],
                           font=("TkDefaultFont", 10, "bold"),
                           borderwidth=1, relief="ridge", padx=5, pady=3)
            lbl.grid(row=0, column=col, sticky="nsew", padx=1, pady=1)
        for col in range(columns):
            lbl = tk.Label(info_table_frame, text=hodnoty[col],
                           font=("TkDefaultFont", 10),
                           borderwidth=1, relief="ridge", padx=5, pady=3)
            lbl.grid(row=1, column=col, sticky="nsew", padx=1, pady=1)
        for col in range(columns):
            info_table_frame.grid_columnconfigure(col, weight=1)

        # VÃ½poÄet svÃ¡tkÅ¯
        selected_year_int = int(selected_year)
        easter_sunday = compute_easter(selected_year_int)
        easter_monday = easter_sunday + timedelta(days=1)
        easter_friday = easter_sunday - timedelta(days=2)
        holidays = [
            date(selected_year_int, 1, 1),
            easter_friday, easter_sunday, easter_monday,
            date(selected_year_int, 5, 1), date(selected_year_int, 5, 8),
            date(selected_year_int, 7, 5), date(selected_year_int, 7, 6),
            date(selected_year_int, 9, 28), date(selected_year_int, 10, 28),
            date(selected_year_int, 11, 17), date(selected_year_int, 12, 25),
            date(selected_year_int, 12, 26)
        ]
        month_frames = {}
        months_info = [
            ("leden", "Leden", 1),
            ("unor", "Ãšnor", 2),
            ("brezen", "BÅ™ezen", 3),
            ("duben", "Duben", 4),
            ("kveten", "KvÄ›ten", 5),
            ("cerven", "ÄŒerven", 6),
            ("cervenec", "ÄŒervenec", 7),
            ("srpen", "Srpen", 8),
            ("zari", "ZÃ¡Å™Ã­", 9),
            ("rijen", "Å˜Ã­jen", 10),
            ("listopad", "Listopad", 11),
            ("prosinec", "Prosinec", 12)
        ]
        current_record_id = record["id"]
        for key, label, month_num in months_info:
            month_frames[key] = render_month_grid(plan_display_frame, selected_year_int, month_num,
                                                    record[key], label, holidays, record["uvazek"], record["smena"],
                                                    editable=True, highlight=False)
        
        half1_hours = 0
        for month in half1_months:
            try:
                day_plan_list = json.loads(record[month])
            except Exception:
                day_plan_list = [""] * 32
            hours, _ = calculate_month_summary(day_plan_list, record["uvazek"])
            half1_hours += hours
        half2_hours = 0
        for month in half2_months:
            try:
                day_plan_list = json.loads(record[month])
            except Exception:
                day_plan_list = [""] * 32
            hours, _ = calculate_month_summary(day_plan_list, record["uvazek"])
            half2_hours += hours
        total_hours = half1_hours + half2_hours
        tk.Label(plan_display_frame, text=f"Leden - ÄŒerven = {half1_hours} hodin", font=("TkDefaultFont", 10, "italic")).pack(pady=5)
        tk.Label(plan_display_frame, text=f"ÄŒervenec - Prosinec = {half2_hours} hodin", font=("TkDefaultFont", 10, "italic")).pack(pady=5)
        tk.Label(plan_display_frame, text=f"Celkem: {total_hours} hodin", font=("TkDefaultFont", 10, "italic", "bold")).pack(pady=5)
        
        if not hasattr(employee_frame, "save_btn_frame"):
            employee_frame.save_btn_frame = tk.Frame(employee_frame)
            employee_frame.save_btn_frame.pack(side=tk.LEFT, padx=5)
        else:
            for widget in employee_frame.save_btn_frame.winfo_children():
                widget.destroy()
        
        btn_state = "normal" if current_user_role in ["velitel", "admin", "superadmin"] else "disabled"
        save_btn = ttk.Button(employee_frame.save_btn_frame, text="UloÅ¾it zmÄ›ny", command=save_employee_plan, state=btn_state)
        save_btn.pack(side=tk.LEFT, padx=5)
        
    except Exception as e:
        logging.error(f"Chyba pÅ™i zobrazenÃ­ plÃ¡nu zamÄ›stnance: {e}")
        messagebox.showerror("Chyba", f"DoÅ¡lo k chybÄ› pÅ™i zobrazenÃ­ plÃ¡nu: {e}")

# ----- Funkce pro uloÅ¾enÃ­ zmÄ›n v plÃ¡nu zamÄ›stnance
def save_employee_plan():
    """
    UloÅ¾Ã­ upravenÃ½ plÃ¡n zamÄ›stnance do databÃ¡ze.
    """
    try:
        if current_record_id is None:
            messagebox.showwarning("UpozornÄ›nÃ­", "NenÃ­ naÄten Å¾Ã¡dnÃ½ plÃ¡n ke uloÅ¾enÃ­.")
            return
        updated_data = {}
        months_info = [
            ("leden", "Leden", 1),
            ("unor", "Ãšnor", 2),
            ("brezen", "BÅ™ezen", 3),
            ("duben", "Duben", 4),
            ("kveten", "KvÄ›ten", 5),
            ("cerven", "ÄŒerven", 6),
            ("cervenec", "ÄŒervenec", 7),
            ("srpen", "Srpen", 8),
            ("zari", "ZÃ¡Å™Ã­", 9),
            ("rijen", "Å˜Ã­jen", 10),
            ("listopad", "Listopad", 11),
            ("prosinec", "Prosinec", 12)
        ]
        for key, label, month_num in months_info:
            frame = month_frames.get(key)
            if frame is not None:
                updated_data[key] = json.dumps(frame.day_plan_list, ensure_ascii=False)
        with sqlite3.connect("service_plans.db") as conn:
            cursor = conn.cursor()
            update_query = """
                UPDATE plans
                SET leden = ?, unor = ?, brezen = ?, duben = ?, kveten = ?,
                    cerven = ?, cervenec = ?, srpen = ?, zari = ?, rijen = ?,
                    listopad = ?, prosinec = ?
                WHERE id = ?
            """
            params = (
                updated_data.get("leden", ""),
                updated_data.get("unor", ""),
                updated_data.get("brezen", ""),
                updated_data.get("duben", ""),
                updated_data.get("kveten", ""),
                updated_data.get("cerven", ""),
                updated_data.get("cervenec", ""),
                updated_data.get("srpen", ""),
                updated_data.get("zari", ""),
                updated_data.get("rijen", ""),
                updated_data.get("listopad", ""),
                updated_data.get("prosinec", ""),
                current_record_id
            )
            cursor.execute(update_query, params)
            conn.commit()
        messagebox.showinfo("ÃšspÄ›ch", "ZmÄ›ny byly ÃºspÄ›Å¡nÄ› uloÅ¾eny.")
        refresh_treeview()
    except Exception as e:
        logging.error(f"Chyba pÅ™i uklÃ¡dÃ¡nÃ­ plÃ¡nu zamÄ›stnance: {e}")
        messagebox.showerror("Chyba", f"DoÅ¡lo k chybÄ› pÅ™i uklÃ¡dÃ¡nÃ­: {e}")

# ----- Funkce pro smazÃ¡nÃ­ plÃ¡nÅ¯ pro zadanÃ½ rok
def delete_plans_by_year():
    """
    SmaÅ¾e vÅ¡echny plÃ¡ny pro zadanÃ½ rok.
    """
    try:
        year_to_delete = simpledialog.askstring("Smazat plÃ¡ny", "Zadejte rok, pro kterÃ½ chcete smazat vÅ¡echny plÃ¡ny:")
        if not year_to_delete:
            return
        if messagebox.askyesno("PotvrzenÃ­", f"Opravdu chcete smazat vÅ¡echny plÃ¡ny pro rok {year_to_delete}?"):
            with sqlite3.connect("service_plans.db") as conn:
                cursor = conn.cursor()
                cursor.execute("DELETE FROM plans WHERE roky = ?", (year_to_delete,))
                conn.commit()
            log_action(f"VÅ¡echny plÃ¡ny pro rok {year_to_delete} byly smazÃ¡ny")
            messagebox.showinfo("ÃšspÄ›ch", f"PlÃ¡ny pro rok {year_to_delete} byly smazÃ¡ny.")
            refresh_treeview()
    except Exception as e:
        logging.error(f"Chyba pÅ™i mazÃ¡nÃ­ plÃ¡nÅ¯ pro rok {year_to_delete}: {e}")
        messagebox.showerror("Chyba", f"DoÅ¡lo k chybÄ› pÅ™i mazÃ¡nÃ­ plÃ¡nÅ¯: {e}")

# ----- Funkce pro zobrazenÃ­ plÃ¡nu smÄ›ny v zÃ¡loÅ¾ce SmÄ›na
def zobraz_plan_smeny():
    """
    ZobrazÃ­ plÃ¡n smÄ›ny podle vybranÃ©ho roku, mÄ›sÃ­ce a smÄ›ny.
    """
    rok = combo_rok_smena.get().strip()
    mesic = combo_mesic_smena.get().strip()
    smena = combo_smena_smena.get().strip()
    if not rok or not mesic or not smena:
        messagebox.showwarning("UpozornÄ›nÃ­", "Vyberte Rok, MÄ›sÃ­c a SmÄ›nu.")
        return
    try:
        rok_int = int(rok)
    except ValueError:
        messagebox.showerror("Chyba", "Rok musÃ­ bÃ½t ÄÃ­slo.")
        return
    mesice = {
        "Leden": ("leden", 1),
        "Ãšnor": ("unor", 2),
        "BÅ™ezen": ("brezen", 3),
        "Duben": ("duben", 4),
        "KvÄ›ten": ("kveten", 5),
        "ÄŒerven": ("cerven", 6),
        "ÄŒervenec": ("cervenec", 7),
        "Srpen": ("srpen", 8),
        "ZÃ¡Å™Ã­": ("zari", 9),
        "Å˜Ã­jen": ("rijen", 10),
        "Listopad": ("listopad", 11),
        "Prosinec": ("prosinec", 12)
    }
    if mesic not in mesice:
        messagebox.showerror("Chyba", "NeznÃ¡mÃ½ mÄ›sÃ­c.")
        return
    col_name, month_num = mesice[mesic]
    with sqlite3.connect("service_plans.db") as conn:
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        query = "SELECT * FROM plans WHERE roky = ? AND smena = ? ORDER BY CAST(poradi as INTEGER) ASC"
        cursor.execute(query, (rok, smena))
        plans = cursor.fetchall()
    if not plans:
        messagebox.showinfo("Informace", "Pro zadanÃ¡ kritÃ©ria nebyl nalezen Å¾Ã¡dnÃ½ plÃ¡n.")
        return
    for widget in smena_display_frame.winfo_children():
        widget.destroy()
    canvas = tk.Canvas(smena_display_frame)
    canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    scrollbar = ttk.Scrollbar(smena_display_frame, orient="vertical", command=canvas.yview)
    scrollbar.pack(side=tk.RIGHT, fill="y")
    canvas.configure(yscrollcommand=scrollbar.set)
    display_frame = tk.Frame(canvas)
    canvas.create_window((0, 0), window=display_frame, anchor="nw")
    def on_frame_configure(event):
        canvas.configure(scrollregion=canvas.bbox("all"))
    display_frame.bind("<Configure>", on_frame_configure)
    easter_sunday = compute_easter(rok_int)
    easter_monday = easter_sunday + timedelta(days=1)
    easter_friday = easter_sunday - timedelta(days=2)
    holidays = [
        date(rok_int, 1, 1),
        easter_friday, easter_sunday, easter_monday,
        date(rok_int, 5, 1), date(rok_int, 5, 8),
        date(rok_int, 7, 5), date(rok_int, 7, 6),
        date(rok_int, 9, 28), date(rok_int, 10, 28),
        date(rok_int, 11, 17), date(rok_int, 12, 25),
        date(rok_int, 12, 26)
    ]
    for plan in plans:
        header_text = f"PoÅ™adÃ­: {plan['poradi']} - {plan['jmeno_prijmeni']}"
        emp_frame = tk.LabelFrame(display_frame, text=header_text, font=("TkDefaultFont", 10, "bold"))
        emp_frame.pack(fill=tk.X, padx=10, pady=5)
        plan_json = plan[col_name]
        # MOD: pÅ™edÃ¡nÃ­ smÄ›ny (plan["smena"]) do renderovÃ¡nÃ­
        render_month_grid(emp_frame, rok_int, month_num, plan_json, mesic, holidays, plan["uvazek"], plan["smena"], editable=False, highlight=False)

# ----- HlavnÃ­ ÄÃ¡st GUI a konfigurace oken
root = tk.Tk()
root.title("SprÃ¡va PlÃ¡nu SluÅ¾eb a DovolenÃ½ch")
root.geometry("1370x900+0+0")

# ----- VytvoÅ™enÃ­ hlavnÃ­ho menu a pÅ™idÃ¡nÃ­ poloÅ¾ky NÃ¡povÄ›da a O Alikaci
menubar = tk.Menu(root)
root.config(menu=menubar)
help_menu = tk.Menu(menubar, tearoff=0)
menubar.add_cascade(label="Menu", menu=help_menu)
help_menu.add_command(label="NÃ¡povÄ›da", command=show_help)
help_menu.add_command(label="O aplikaci", command=show_verze)

# ----- PÅ™ihlaÅ¡ovacÃ­ panel
login_frame = tk.Frame(root)
login_frame.pack(fill=tk.X, padx=10, pady=5)
tk.Label(login_frame, text="JmÃ©no:").pack(side=tk.LEFT, padx=5)
login_name_entry = tk.Entry(login_frame)
login_name_entry.pack(side=tk.LEFT, padx=5)
tk.Label(login_frame, text="Heslo:").pack(side=tk.LEFT, padx=5)
login_pwd_entry = tk.Entry(login_frame, show="*")
login_pwd_entry.pack(side=tk.LEFT, padx=5)
login_button = tk.Button(login_frame, text="PÅ™ihlÃ¡sit se", command=login)
login_button.pack(side=tk.LEFT, padx=5)
logout_button = tk.Button(login_frame, text="OdhlÃ¡sit se", command=logout)
logout_button.pack(side=tk.LEFT, padx=5)
login_status_label = tk.Label(login_frame, text="Nejste pÅ™ihlÃ¡Å¡eni: MÃ¡te status (uÅ¾ivatel)")
login_status_label.pack(side=tk.LEFT, padx=10)

# ----- HlavnÃ­ zÃ¡loÅ¾kovÃ½ widget
notebook = ttk.Notebook(root)
notebook.pack(expand=True, fill='both')

# ZÃ¡loÅ¾ka PlÃ¡ny
tab_plany = ttk.Frame(notebook, width=1200, height=700)
tab_plany.pack(fill=tk.BOTH, expand=True)
notebook.add(tab_plany, text="PlÃ¡ny")

# PodzÃ¡loÅ¾ky pro PlÃ¡ny: ZamÄ›stnanec a SmÄ›na
plans_notebook = ttk.Notebook(tab_plany)
plans_notebook.pack(expand=True, fill='both', padx=10, pady=10)
tab_zamestnanec = ttk.Frame(plans_notebook)
plans_notebook.add(tab_zamestnanec, text="ZamÄ›stnanec")
tab_smena = ttk.Frame(plans_notebook)
plans_notebook.add(tab_smena, text="SmÄ›na")

# ----- Filtr pro zÃ¡loÅ¾ku SmÄ›na
filter_smena_frame = tk.Frame(tab_smena)
filter_smena_frame.pack(fill=tk.X, padx=10, pady=5)
current_year = datetime.now().year
years = [str(current_year - 1), str(current_year), str(current_year + 1)]
months = ["Leden", "Ãšnor", "BÅ™ezen", "Duben", "KvÄ›ten", "ÄŒerven", "ÄŒervenec", "Srpen", "ZÃ¡Å™Ã­", "Å˜Ã­jen", "Listopad", "Prosinec"]
shifts = ["", "SmÄ›na 1", "SmÄ›na 2", "SmÄ›na 3", "SmÄ›na 4", "SmÄ›na 5", "SmÄ›na 6"]

tk.Label(filter_smena_frame, text="Roky:", font=("TkDefaultFont", 10, "bold")).pack(side=tk.LEFT, padx=5)
combo_rok_smena = ttk.Combobox(filter_smena_frame, values=years, state="readonly", width=10)
combo_rok_smena.current(1)
combo_rok_smena.pack(side=tk.LEFT, padx=5)
tk.Label(filter_smena_frame, text="MÄ›sÃ­c:", font=("TkDefaultFont", 10, "bold")).pack(side=tk.LEFT, padx=5)
combo_mesic_smena = ttk.Combobox(filter_smena_frame, values=months, state="readonly", width=10)
combo_mesic_smena.current(datetime.now().month - 1)
combo_mesic_smena.pack(side=tk.LEFT, padx=5)
tk.Label(filter_smena_frame, text="SmÄ›na:", font=("TkDefaultFont", 10, "bold")).pack(side=tk.LEFT, padx=5)
combo_smena_smena = ttk.Combobox(filter_smena_frame, state="readonly", width=10)
combo_smena_smena['values'] = ["", "SmÄ›na 1", "SmÄ›na 2", "SmÄ›na 3", "SmÄ›na 4", "SmÄ›na 5", "SmÄ›na 6"]
combo_smena_smena.set("")
combo_smena_smena.pack(side=tk.LEFT, padx=5)
btn_zobraz_plan_smeny_smena = ttk.Button(filter_smena_frame, text="Zobraz plÃ¡n SmÄ›ny", command=zobraz_plan_smeny)
btn_zobraz_plan_smeny_smena.pack(side=tk.LEFT, padx=5)

# MOD: PÅ™idÃ¡nÃ­ tlaÄÃ­tek Zamknout a Odemknout plÃ¡n â€“ ty budou viditelnÃ© pouze pro superadmina
btn_lock_plan = ttk.Button(filter_smena_frame, text="Zamknout PlÃ¡n", command=lock_plan)
btn_lock_plan.pack(side=tk.LEFT, padx=5)
btn_unlock_plan = ttk.Button(filter_smena_frame, text="Odemknout PlÃ¡n", command=unlock_plan)
btn_unlock_plan.pack(side=tk.LEFT, padx=5)

smena_display_frame = tk.Frame(tab_smena)
smena_display_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

# ----- ZÃ¡loÅ¾ka NastavenÃ­
tab_nastaveni = ttk.Frame(notebook)
notebook.add(tab_nastaveni, text="NastavenÃ­")

filter_frame = tk.Frame(tab_nastaveni)
filter_frame.pack(pady=10, padx=10, anchor="w")
tk.Label(filter_frame, text="Filtr - JmÃ©no a pÅ™Ã­jmenÃ­:").pack(side=tk.LEFT, padx=5)
filter_jmeno = ttk.Combobox(filter_frame, state="readonly", width=20)
filter_jmeno.pack(side=tk.LEFT, padx=5)
tk.Label(filter_frame, text="Rok:").pack(side=tk.LEFT, padx=5)
filter_rok = ttk.Combobox(filter_frame, state="readonly", width=10)
filter_rok.pack(side=tk.LEFT, padx=5)
tk.Label(filter_frame, text="SmÄ›na:").pack(side=tk.LEFT, padx=5)
filter_smena = ttk.Combobox(filter_frame, state="readonly", width=15)
filter_smena.pack(side=tk.LEFT, padx=5)

def apply_filters():
    selected_jmeno = filter_jmeno.get()
    selected_rok = filter_rok.get()
    selected_smena = filter_smena.get()
    refresh_treeview_filtered(jmeno_filter=selected_jmeno, rok_filter=selected_rok, smena_filter=selected_smena)

tk.Button(filter_frame, text="Filtrovat", command=apply_filters).pack(side=tk.LEFT, padx=5)
tk.Button(filter_frame, text="Zobrazit vÅ¡e", command=refresh_treeview).pack(side=tk.LEFT, padx=5)

style = ttk.Style()
style.configure("Treeview.Heading", font=("TkDefaultFont", 10, "bold"))

tree = ttk.Treeview(tab_nastaveni, columns=("jmeno_prijmeni", "smena", "uvazek", "roky"), show="headings")
tree.heading("jmeno_prijmeni", text="JmÃ©no a pÅ™Ã­jmenÃ­", command=lambda: treeview_sort_column(tree, "jmeno_prijmeni", False))
tree.heading("smena", text="SmÄ›na", command=lambda: treeview_sort_column(tree, "smena", False))
tree.heading("uvazek", text="Ãšvazek", command=lambda: treeview_sort_column(tree, "uvazek", False))
tree.heading("roky", text="Rok", command=lambda: treeview_sort_column(tree, "roky", False))
tree.column("smena", width=90)
tree.column("uvazek", width=100)
tree.column("roky", width=60)
tree.pack(pady=20, padx=20, fill=tk.BOTH, expand=True)

# PÅ™idÃ¡me binding pro posouvÃ¡nÃ­ treeview koleÄkem myÅ¡i
tree.bind("<MouseWheel>", lambda event: tree.yview_scroll(-1 * (event.delta // 120), "units"))

button_frame = ttk.Frame(tab_nastaveni)
button_frame.pack(pady=10)
btn_new = ttk.Button(button_frame, text="NovÃ½", command=on_new)
btn_new.grid(row=0, column=0, padx=5)
btn_edit = ttk.Button(button_frame, text="Edit", command=on_edit)
btn_edit.grid(row=0, column=1, padx=5)
btn_delete = ttk.Button(button_frame, text="Smazat", command=on_delete)
btn_delete.grid(row=0, column=2, padx=5)
btn_data = ttk.Button(button_frame, text="Data", command=open_settings_window)
btn_data.grid(row=0, column=3, padx=5)
# PÅ™idÃ¡nÃ­ tlaÄÃ­tka "Fond" za tlaÄÃ­tkem "Data":
fond_button = ttk.Button(button_frame, text="Fond", command=open_fond_window)
fond_button.grid(row=0, column=4, padx=5)
btn_delete_year = ttk.Button(button_frame, text="Smazat plÃ¡ny pro rok", command=delete_plans_by_year)
btn_delete_year.grid(row=0, column=5, padx=5)

notebook.bind("<<NotebookTabChanged>>", lambda event: refresh_treeview() if event.widget.tab(event.widget.index("current"), "text") == "NastavenÃ­" else None)

employee_frame = tk.Frame(tab_zamestnanec)
employee_frame.pack(side=tk.TOP, fill=tk.X, padx=10, pady=10)
tk.Label(employee_frame, text="ZamÄ›stnanec:", font=("TkDefaultFont", 10, "bold")).pack(side=tk.LEFT, padx=5)
employee_combobox = ttk.Combobox(employee_frame, state="readonly")
employee_combobox.pack(side=tk.LEFT, padx=5)
tk.Label(employee_frame, text="Rok:", font=("TkDefaultFont", 10, "bold")).pack(side=tk.LEFT, padx=5)
year_combobox = ttk.Combobox(employee_frame, state="readonly")
year_combobox.pack(side=tk.LEFT, padx=5)
tk.Label(employee_frame, text="SmÄ›na:", font=("TkDefaultFont", 10, "bold")).pack(side=tk.LEFT, padx=5)
shift_filter_combobox = ttk.Combobox(employee_frame, state="readonly", width=10)
shift_filter_combobox['values'] = ["", "SmÄ›na 1", "SmÄ›na 2", "SmÄ›na 3", "SmÄ›na 4", "SmÄ›na 5", "SmÄ›na 6"]
shift_filter_combobox.set("")
shift_filter_combobox.pack(side=tk.LEFT, padx=5)
shift_filter_combobox.bind("<<ComboboxSelected>>", lambda event: update_employee_list())
show_plan_button = ttk.Button(employee_frame, text="Zobrazit plÃ¡n", command=show_employee_plan)
show_plan_button.pack(side=tk.LEFT, padx=5)

employee_plan_frame = tk.Frame(tab_zamestnanec)
employee_plan_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

plan_display_canvas = tk.Canvas(employee_plan_frame)
plan_display_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
plan_v_scrollbar = ttk.Scrollbar(employee_plan_frame, orient="vertical", command=plan_display_canvas.yview)
plan_v_scrollbar.pack(side=tk.RIGHT, fill="y")
plan_h_scrollbar = ttk.Scrollbar(employee_plan_frame, orient="horizontal", command=plan_display_canvas.xview)
plan_h_scrollbar.pack(side=tk.BOTTOM, fill="x")
plan_display_canvas.configure(yscrollcommand=plan_v_scrollbar.set, xscrollcommand=plan_h_scrollbar.set)
plan_display_frame = tk.Frame(plan_display_canvas)
plan_display_canvas.create_window((0,0), window=plan_display_frame, anchor="nw")
plan_display_frame.bind("<Configure>", lambda event: plan_display_canvas.configure(scrollregion=plan_display_canvas.bbox("all")))

# ----- Inicializace databÃ¡ze a nastavenÃ­ GUI
init_db()
populate_employee_and_year()
apply_access_control()
refresh_treeview()

# ----- PlÃ¡novanÃ¡ zÃ¡loha databÃ¡ze kaÅ¾dÃ½ch 24 hodin (86400000 ms)
backup_interval_ms = 86400000  # 24 hodin

def schedule_backup():
    """
    PlÃ¡nuje pravidelnou zÃ¡lohu databÃ¡ze.
    """
    backup_dir = "db_backups"
    if os.path.exists(backup_dir):
        backup_files = [os.path.join(backup_dir, f) for f in os.listdir(backup_dir)
                        if f.startswith("service_plans_backup_") and f.endswith(".db")]
        if backup_files:
            latest_backup = max(backup_files, key=os.path.getmtime)
            last_backup_time = os.path.getmtime(latest_backup)
            if time.time() - last_backup_time >= backup_interval_ms / 1000:
                backup_database()
        else:
            backup_database()
    else:
        backup_database()
    root.after(backup_interval_ms, schedule_backup)
    
schedule_backup()

# ----- HlavnÃ­ smyÄka GUI
root.mainloop()
